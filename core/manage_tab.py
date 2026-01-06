from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QLineEdit,
                               QComboBox, QGroupBox, QFileDialog, QApplication)
from PySide6.QtCore import Qt
import os
import json

from core.config import config, get_resource_path
from core.styles import get_font_gray_style, get_button_style

# 导入爬虫相关的类
from .achievement_table import AchievementTable


def _standardize_achievement_fields(achievement):
    """标准化成就字段名和值"""
    # 标准化字段名映射
    field_mapping = {
        'version': '版本',
        'name': '名称',
        'description': '描述',
        'reward': '奖励',
        'is_hidden': '是否隐藏',
        'status': '获取状态'
    }
    
    # 应用字段名映射
    standardized = {}
    for key, value in achievement.items():
        mapped_key = field_mapping.get(key, key)
        standardized[mapped_key] = value
    
    # 处理版本字段格式
    if '版本' in standardized:
        version = standardized['版本']
        # 如果是整数形式的字符串，转换为浮点数格式
        try:
            if version and '.' not in version:
                # 检查是否为纯数字
                if version.isdigit():
                    standardized['版本'] = f"{int(version)}.0"
                else:
                    # 尝试转换为浮点数
                    float_val = float(version)
                    standardized['版本'] = f"{float_val}.0" if float_val == int(float_val) else str(float_val)
        except (ValueError, TypeError):
            # 如果转换失败，保持原值
            pass
    
    # 处理获取状态字段格式
    if '获取状态' in standardized:
        status = standardized['获取状态']
        # 兼容旧的状态格式
        status_mapping = {
            'completed': '已完成',
            '': '未完成',
            'unavailable': '暂不可获取'
        }
        standardized['获取状态'] = status_mapping.get(status, status)
    
    # 处理是否隐藏字段
    if '是否隐藏' in standardized:
        is_hidden = standardized['是否隐藏']
        if isinstance(is_hidden, bool):
            # 如果已经是布尔值，转换为字符串
            standardized['是否隐藏'] = '隐藏' if is_hidden else ''
        elif is_hidden not in ['隐藏', '']:
            # 如果值不是期望的字符串，进行转换
            standardized['是否隐藏'] = '隐藏' if is_hidden else ''
    
    return standardized




class AchievementManager:
    """成就管理器"""
    
    def __init__(self):
        self.achievements = []
        self.filtered_achievements = []
        
    def load_data(self, achievements):
        """加载数据"""
        self.achievements = achievements
        self.filtered_achievements = achievements.copy()
        
    def filter_data(self, search_text="", version="", first_category="", second_category="", 
                   hidden_type="all", priority="默认排序", obtainable="全部"):
        """筛选数据"""
        self.filtered_achievements = []
        
        for achievement in self.achievements:
            # 搜索文本
            if search_text and search_text.lower() not in achievement.get('名称', '').lower() and \
               search_text.lower() not in achievement.get('描述', '').lower():
                continue
            
            # 版本筛选
            if version and version != "所有版本" and achievement.get('版本', '') != version:
                continue
            
            # 第一分类筛选
            if first_category and first_category != "全部" and achievement.get('第一分类', '') != first_category:
                continue
            
            # 第二分类筛选
            if second_category and second_category != "全部" and achievement.get('第二分类', '') != second_category:
                continue
            
            # 隐藏状态筛选
            if hidden_type == "hidden_only" and achievement.get('是否隐藏') != '隐藏':
                continue
            elif hidden_type == "not_hidden" and achievement.get('是否隐藏') == '隐藏':
                continue
            
            # 获取类型筛选
            if obtainable == "可获取" and achievement.get('获取状态', '') == "暂不可获取":
                continue
            elif obtainable == "暂不可获取" and achievement.get('获取状态', '') != "暂不可获取":
                continue
            elif obtainable == "多选一":
                # 只显示成就组，显示所有组成员
                group_id = achievement.get('成就组ID')
                if not group_id:
                    # 非成就组的成就不显示
                    continue
            
            self.filtered_achievements.append(achievement)
        
        # 排序处理
        if priority == "未完成优先":
            # 将未完成的排前面
            self.filtered_achievements.sort(key=lambda x: (
                0 if x.get('获取状态', '') in ['', '未完成'] else 1,
                int(x.get('绝对编号', '0'))  # 使用绝对编号排序
            ))
        else:
            # 默认按绝对编号排序（绝对编号仅用于排序）
            self.filtered_achievements.sort(key=lambda x: int(x.get('绝对编号', '0')))
        
        return self.filtered_achievements
    
    def get_statistics(self):
        """获取统计信息"""
        # 正确统计总计（考虑成就组）
        total_groups = set()
        total_achievements = 0
        for achievement in self.achievements:
            group_id = achievement.get('成就组ID')
            if group_id:
                total_groups.add(group_id)
            else:
                total_achievements += 1
        total = len(total_groups) + total_achievements
        
        # 正确统计完成数（考虑成就组）
        completed = 0
        processed_groups = set()
        for achievement in self.achievements:
            status = achievement.get('获取状态', '')
            group_id = achievement.get('成就组ID')
            
            if status == '已完成':
                if group_id:
                    # 成就组：只计算一次
                    if group_id not in processed_groups:
                        completed += 1
                        processed_groups.add(group_id)
                else:
                    # 普通成就
                    completed += 1
        
        hidden = sum(1 for a in self.achievements if a.get('是否隐藏') == '隐藏')
        
        return {
            'total': total,
            'completed': completed,
            'hidden': hidden
        }


class ManageTab(QWidget):
    """成就管理标签页"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("manage_tab")
        self.manager = AchievementManager()
        self.init_ui()
        
        # 监听用户切换信号
        from core.signal_bus import signal_bus
        signal_bus.user_switched.connect(self.on_user_switched)
        signal_bus.theme_changed.connect(self.on_theme_changed)
        
        # 尝试加载本地数据
        self.load_local_data()
    
    def on_user_switched(self, username):
        """用户切换时重新加载数据"""
        print(f"[INFO] 用户切换到: {username}")
        self.load_local_data()
    
    def on_theme_changed(self, theme):
        """主题切换时更新按钮样式"""
        from core.styles import get_button_style
        self.import_btn.setStyleSheet(get_button_style(theme))
        self.import_excel_btn.setStyleSheet(get_button_style(theme))
        self.export_excel_btn.setStyleSheet(get_button_style(theme))
        self.export_json_btn.setStyleSheet(get_button_style(theme))
        self.settings_btn.setStyleSheet(get_button_style(theme))
        self.help_btn.setStyleSheet(get_button_style(theme))
        
    def init_ui(self):
        """初始化UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(15)
        
        # 筛选面板
        filter_group = QGroupBox()
        filter_main_layout = QVBoxLayout(filter_group)
        
        # 第一行筛选
        filter_layout = QHBoxLayout()
        
        # 第二行筛选
        filter_layout2 = QHBoxLayout()
        
        # 第一行筛选
        filter_layout.setSpacing(5)
        
        # 成就搜索
        search_label = QLabel("成就搜索:")
        search_label.setFixedWidth(70)
        search_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        filter_layout.addWidget(search_label)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("搜索成就名称或描述...")
        self.search_input.setFixedWidth(270)
        self.search_input.setFixedHeight(26)
        self.search_input.textChanged.connect(self.filter_data)
        filter_layout.addWidget(self.search_input)
        
        # 版本筛选
        version_label = QLabel("版本:")
        version_label.setFixedWidth(70)
        version_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        filter_layout.addWidget(version_label)
        
        self.version_filter = QComboBox()
        self.version_filter.addItem("所有版本")
        self.version_filter.setFixedWidth(100)
        self.version_filter.currentTextChanged.connect(self.filter_data)
        filter_layout.addWidget(self.version_filter)
        
        # 获取类型筛选
        obtainable_label = QLabel("获取类型:")
        obtainable_label.setFixedWidth(70)
        obtainable_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        filter_layout.addWidget(obtainable_label)
        
        self.obtainable_filter = QComboBox()
        self.obtainable_filter.addItem("全部")
        self.obtainable_filter.addItem("可获取")
        self.obtainable_filter.addItem("暂不可获取")
        self.obtainable_filter.addItem("多选一")
        self.obtainable_filter.setFixedWidth(100)
        self.obtainable_filter.currentTextChanged.connect(self.filter_data)
        filter_layout.addWidget(self.obtainable_filter)
        
        filter_layout.addStretch()
        
        # 第二行筛选：第一分类 + 第二分类 + 排序 + 是否隐藏
        filter_layout2 = QHBoxLayout()
        filter_layout2.setSpacing(5)
        
        # 第一分类筛选
        first_category_label = QLabel("第一分类:")
        first_category_label.setFixedWidth(70)
        first_category_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        filter_layout2.addWidget(first_category_label)
        
        self.first_category_filter = QComboBox()
        self.first_category_filter.addItem("全部")
        self.first_category_filter.setFixedWidth(100)
        self.first_category_filter.currentTextChanged.connect(self.on_first_category_changed)
        filter_layout2.addWidget(self.first_category_filter)
        
        # 第二分类筛选
        second_category_label = QLabel("第二分类:")
        second_category_label.setFixedWidth(70)
        second_category_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        filter_layout2.addWidget(second_category_label)
        
        self.second_category_filter = QComboBox()
        self.second_category_filter.addItem("全部")
        self.second_category_filter.setFixedWidth(100)
        self.second_category_filter.currentTextChanged.connect(self.filter_data)
        filter_layout2.addWidget(self.second_category_filter)
        
        # 排序筛选
        priority_label = QLabel("排序:")
        priority_label.setFixedWidth(70)
        priority_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        filter_layout2.addWidget(priority_label)
        
        self.priority_filter = QComboBox()
        self.priority_filter.addItem("默认排序")
        self.priority_filter.addItem("未完成优先")
        self.priority_filter.setFixedWidth(100)
        self.priority_filter.currentTextChanged.connect(self.filter_data)
        filter_layout2.addWidget(self.priority_filter)
        
        # 是否隐藏筛选
        hidden_label = QLabel("是否隐藏:")
        hidden_label.setFixedWidth(70)
        hidden_label.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        filter_layout2.addWidget(hidden_label)
        
        self.hidden_filter = QComboBox()
        self.hidden_filter.addItem("所有")
        self.hidden_filter.addItem("仅隐藏")
        self.hidden_filter.addItem("仅非隐藏")
        self.hidden_filter.setFixedWidth(100)
        self.hidden_filter.currentTextChanged.connect(self.filter_data)
        filter_layout2.addWidget(self.hidden_filter)
        
        filter_layout2.addStretch()
        
        # 将两行布局添加到筛选主布局中
        filter_main_layout.addLayout(filter_layout)
        filter_main_layout.addLayout(filter_layout2)
        
        # 减少筛选区域的垂直间距
        filter_main_layout.setSpacing(3)
        filter_main_layout.setContentsMargins(8, 5, 8, 5)
        
        # 按钮区域
        button_group = QGroupBox()
        button_layout = QVBoxLayout(button_group)
        button_layout.setSpacing(5)
        button_layout.setContentsMargins(8, 5, 8, 5)
        
        # 第一行：导入JSON、导出JSON、设置
        first_row = QHBoxLayout()
        first_row.setSpacing(5)
        self.import_btn = QPushButton("导入JSON")
        self.import_btn.setStyleSheet(get_button_style(config.theme))
        self.import_btn.clicked.connect(self.import_json)
        self.import_btn.setFixedWidth(80)
        first_row.addWidget(self.import_btn)
        
        self.export_json_btn = QPushButton("导出JSON")
        self.export_json_btn.setStyleSheet(get_button_style(config.theme))
        self.export_json_btn.clicked.connect(self.export_full_json)
        self.export_json_btn.setFixedWidth(80)
        first_row.addWidget(self.export_json_btn)
        
        self.settings_btn = QPushButton("设置")
        self.settings_btn.setStyleSheet(get_button_style(config.theme))
        self.settings_btn.clicked.connect(self.open_settings)
        self.settings_btn.setFixedWidth(80)
        first_row.addWidget(self.settings_btn)
        first_row.addStretch()
        
        button_layout.addLayout(first_row)
        
        # 第二行：导入Excel、导出Excel、帮助
        second_row = QHBoxLayout()
        second_row.setSpacing(5)
        self.import_excel_btn = QPushButton("导入Excel")
        self.import_excel_btn.setStyleSheet(get_button_style(config.theme))
        self.import_excel_btn.clicked.connect(self.import_excel)
        self.import_excel_btn.setFixedWidth(80)
        second_row.addWidget(self.import_excel_btn)
        
        self.export_excel_btn = QPushButton("导出Excel")
        self.export_excel_btn.setStyleSheet(get_button_style(config.theme))
        self.export_excel_btn.clicked.connect(self.export_excel)
        self.export_excel_btn.setFixedWidth(80)
        second_row.addWidget(self.export_excel_btn)
        
        self.help_btn = QPushButton("帮助")
        self.help_btn.setStyleSheet(get_button_style(config.theme))
        self.help_btn.clicked.connect(self.open_help)
        self.help_btn.setFixedWidth(80)
        second_row.addWidget(self.help_btn)
        second_row.addStretch()
        
        button_layout.addLayout(second_row)
        
        # 创建主水平布局，包含筛选区域和按钮区域
        main_filter_layout = QHBoxLayout()
        main_filter_layout.addWidget(filter_group, 3)  # 筛选区域占3份
        main_filter_layout.addWidget(button_group, 1)   # 按钮区域占1份
        
        layout.addLayout(main_filter_layout)
        
        # 统计信息
        stats_group = QGroupBox("统计信息")
        stats_layout = QHBoxLayout(stats_group)
        
        # 总计
        self.total_label = QLabel("📊 总计: 0")
        stats_layout.addWidget(self.total_label)
        
        # 已完成
        self.completed_label = QLabel("✅ 已完成: 0")
        stats_layout.addWidget(self.completed_label)
        
        # 未完成
        self.incomplete_label = QLabel("⭕ 未完成: 0")
        stats_layout.addWidget(self.incomplete_label)
        
        # 隐藏成就
        self.hidden_label = QLabel("🙈 隐藏成就: 0")
        stats_layout.addWidget(self.hidden_label)
        
        # 暂不可获取
        self.unavailable_label = QLabel("🚫 暂不可获取: 0")
        stats_layout.addWidget(self.unavailable_label)
        
        # 多选一
        self.multi_choice_label = QLabel("🎯 多选一: 0")
        stats_layout.addWidget(self.multi_choice_label)
        
        stats_layout.addStretch()
        layout.addWidget(stats_group)
        
        # 管理表格
        self.manager_table = AchievementTable()
        layout.addWidget(self.manager_table)
        
        # 初始日志
        print("[INFO] 成就管理标签页已初始化")
    
    def load_data(self, achievements):
        """加载数据"""
        # 调试信息：检查数据格式
        if achievements:
            print(f"[DEBUG] 接收到的数据示例 - 是否隐藏字段: {achievements[0].get('是否隐藏', 'None')}")
            print(f"[DEBUG] 接收到的数据示例 - is_hidden字段: {achievements[0].get('is_hidden', 'None')}")
        
        self.manager.load_data(achievements)
        
        # 更新版本筛选器
        versions = set()
        categories = set()
        
        for achievement in achievements:
            versions.add(achievement.get('版本', ''))
            categories.add(achievement.get('第一分类', ''))
            categories.add(achievement.get('第二分类', ''))
        
        # 更新版本下拉框
        self.version_filter.clear()
        self.version_filter.addItem("所有版本")
        # 将版本号转换为浮点数进行倒序排序
        valid_versions = []
        
        for v in versions:
            if v:
                try:
                    # 尝试转换为浮点数进行排序
                    float_val = float(v)
                    valid_versions.append((float_val, v))
                except (ValueError, TypeError):
                    # 如果转换失败，使用原始字符串排序
                    valid_versions.append((-1, v))
        
        # 按浮点数值倒序排序，无法转换的放在最后
        sorted_versions = [v[1] for v in sorted(valid_versions, key=lambda x: x[0], reverse=True)]
        
        for version in sorted_versions:
            self.version_filter.addItem(version)
        
        # 更新分类下拉框
        self.category_filter.clear()
        self.category_filter.addItem("所有分类")
        for category in sorted(categories):
            if category:
                self.category_filter.addItem(category)
        
        # 初始显示所有数据
        self.filter_data()
        
        # 更新统计信息
        self.update_statistics()
        
        # 更新表格显示
        filtered_data = self.manager.filter_data(
            self.search_input.text(),
            self.version_filter.currentText(),
            self.first_category_filter.currentText(),
            self.second_category_filter.currentText(),
            self.hidden_filter.currentText(),
            self.priority_filter.currentText(),
            self.obtainable_filter.currentText()
        )
        self.manager_table.load_data(filtered_data)
        
        print(f"[SUCCESS] 已加载 {len(achievements)} 条成就数据")
    
    def on_first_category_changed(self):
        """第一分类变化时更新第二分类选项并重置为全部"""
        first_category = self.first_category_filter.currentText()
        
        # 清空第二分类选项
        self.second_category_filter.clear()
        self.second_category_filter.addItem("全部")
        
        # 获取分类配置
        category_config = config.load_category_config()
        
        # 根据第一分类筛选第二分类
        if first_category != "全部":
            second_categories = set()
            for achievement in self.manager.achievements:
                if achievement.get('第一分类', '') == first_category:
                    second_categories.add(achievement.get('第二分类', ''))
            
            # 按照配置中的顺序添加第二分类
            if first_category in category_config.get("second_categories", {}):
                second_category_order = category_config["second_categories"][first_category]
                ordered_second_categories = sorted(second_categories, key=lambda x: int(second_category_order.get(x, 999)))
                for category in ordered_second_categories:
                    if category:
                        self.second_category_filter.addItem(category)
            else:
                # 如果没有配置，按字母顺序
                for category in sorted(second_categories):
                    if category:
                        self.second_category_filter.addItem(category)
        else:
            # 显示所有第二分类
            second_categories = set()
            for achievement in self.manager.achievements:
                second_categories.add(achievement.get('第二分类', ''))
            
            for category in sorted(second_categories):
                if category:
                    self.second_category_filter.addItem(category)
        
        # 强制重置第二分类为"全部"
        self.second_category_filter.setCurrentIndex(0)
        
        # 触发筛选
        self.filter_data()
    
    def filter_data(self):
        """筛选数据"""
        search_text = self.search_input.text().strip()
        version = self.version_filter.currentText()
        first_category = self.first_category_filter.currentText()
        second_category = self.second_category_filter.currentText()
        hidden_type = self.hidden_filter.currentText()
        priority = self.priority_filter.currentText()
        obtainable = self.obtainable_filter.currentText()
        
        # 映射隐藏类型
        hidden_map = {
            "所有": "all",
            "仅隐藏": "hidden_only",
            "仅非隐藏": "not_hidden"
        }
        hidden_type = hidden_map.get(hidden_type, "all")
        
        # 筛选数据
        filtered = self.manager.filter_data(
            search_text, version, first_category, second_category, 
            hidden_type, priority, obtainable
        )
        
        # 在多选一模式下，为每个成就添加组标识
        if obtainable == "多选一":
            # 为筛选后的成就添加组标识（保存原始名称）
            for achievement in filtered:
                group_id = achievement.get('成就组ID')
                if group_id:
                    # 简单的组标识
                    group_number = group_id.split('_')[1] if '_' in group_id else '1'
                    original_name = achievement.get('原始名称', '') or achievement.get('名称', '')
                    # 保存原始名称
                    achievement['原始名称'] = original_name
                    # 修改显示名称
                    achievement['名称'] = f"[{group_number}] {original_name}"
        else:
            # 恢复原始名称
            for achievement in filtered:
                if '原始名称' in achievement:
                    achievement['名称'] = achievement['原始名称']
        
        # 更新表格
        self.manager_table.load_data(filtered)
        
        # 更新统计信息
        self.update_statistics(filtered)
    
    def update_statistics(self, data=None):
        """更新统计信息"""
        if data is None:
            data = self.manager.filtered_achievements
        
        # 使用统一的统计方法
        statistics = self.calculate_statistics(data)
        
        self.total_label.setText(f"📊 总计: {statistics['total']}")
        self.completed_label.setText(f"✅ 已完成: {statistics['completed']}")
        self.incomplete_label.setText(f"⭕ 未完成: {statistics['incomplete']}")
        self.hidden_label.setText(f"🙈 隐藏成就: {statistics['hidden']}")
        self.unavailable_label.setText(f"🚫 暂不可获取: {statistics['unavailable']}")
        self.multi_choice_label.setText(f"🎯 多选一成就: {statistics['multi_choice']}")
    
    def open_settings(self):
        """打开设置对话框"""
        from core.settings_dialog import TemplateSettingsDialog
        dialog = TemplateSettingsDialog(self)
        dialog.exec()
    
    def open_help(self):
        """打开帮助对话框"""
        from core.help_dialog import HelpDialog
        dialog = HelpDialog(self)
        dialog.exec()
    
    def import_excel(self):
        """导入Excel文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "导入Excel文件", "", "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            try:
                # 显示确认对话框
                from core.custom_message_box import CustomMessageBox
                reply = CustomMessageBox.question(
                    self, 
                    "确认导入", 
                    "导入Excel将会覆盖当前的成就数据和用户进度数据，此操作不可撤销！\n\n确定要继续吗？",
                    ("确定", "取消")
                )
                
                if reply != CustomMessageBox.Yes:
                    print("[INFO] 用户取消了Excel导入操作")
                    return
                
                self.import_from_excel(file_path)
            except Exception as e:
                print(f"[ERROR] 导入Excel失败: {str(e)}")
                show_notification(self, f"导入Excel失败: {str(e)}")
    
    def import_from_excel(self, excel_path):
        """从Excel文件导入数据并转换为base_achievements和user_progress"""
        try:
            from openpyxl import load_workbook
            
            # 读取Excel文件
            print(f"[INFO] 正在读取Excel文件: {excel_path}")
            workbook = load_workbook(excel_path)
            sheet = workbook.active
            
            # 检查第一行是否为时间信息行
            first_row_values = [cell.value for cell in sheet[1]]
            is_info_row = any("导出时间:" in str(val) or "用户:" in str(val) for val in first_row_values if val)
            
            # 根据第一行内容确定表头位置
            if is_info_row:
                # 第一行是信息行，表头在第二行
                header_row = 2
                data_start_row = 3
            else:
                # 第一行就是表头
                header_row = 1
                data_start_row = 2
            
            # 获取表头
            headers = []
            for cell in sheet[header_row]:
                headers.append(cell.value)
            
            # 检查必要的列
            required_columns = ['名称', '第二分类']
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                raise Exception(f"缺少必要的列: {', '.join(missing_columns)}")
            
            # 创建列名到索引的映射
            col_index = {header: idx for idx, header in enumerate(headers)}
            
            # 数据清洗和转换
            print(f"[INFO] 开始数据清洗...")
            print(f"[DEBUG] 总行数: {sheet.max_row}, 数据开始行: {data_start_row}")
            print(f"[DEBUG] 列索引: {col_index}")
            achievements = []
            
            # 加载分类配置
            try:
                # 确保config对象已正确初始化
                if not hasattr(config, 'load_category_config'):
                    print("[ERROR] config对象缺少load_category_config方法")
                    raise Exception("配置对象未正确初始化")
                
                category_config = config.load_category_config()
                if not isinstance(category_config, dict):
                    print(f"[ERROR] category_config不是字典类型: {type(category_config)}")
                    category_config = {}
                
                first_categories = category_config.get("first_categories", {})
                second_categories = category_config.get("second_categories", {})
                
                # 确保返回的是字典
                if not isinstance(first_categories, dict):
                    first_categories = {}
                if not isinstance(second_categories, dict):
                    second_categories = {}
                    
            except Exception as e:
                print(f"[ERROR] 加载分类配置失败: {str(e)}")
                # 使用默认配置
                first_categories = {}
                second_categories = {}
                print("[INFO] 使用默认分类配置")
            
            # 创建第二分类到第一分类的映射
            first_category_map = {}
            for first_cat, second_cats in second_categories.items():
                for second_cat in second_cats:
                    first_category_map[second_cat] = first_cat
            
            # 预先扫描所有第二分类，检查是否有缺失的
            missing_second_categories = set()
            for row_idx, row in enumerate(sheet.iter_rows(min_row=data_start_row), start=data_start_row):
                if not any(cell.value for cell in row):
                    continue  # 跳过空行
                
                # 获取第二分类
                if '第二分类' in col_index:
                    second_category_value = row[col_index['第二分类']].value
                    second_category = str(second_category_value).strip() if second_category_value else ''
                    
                    if row_idx <= data_start_row + 2:  # 只显示前几行的调试信息
                        print(f"[DEBUG] 预扫描第{row_idx}行，第二分类: '{second_category}'")
                        print(f"[DEBUG] 预扫描第{row_idx}行，是否在first_category_map中: {second_category in first_category_map}")
                    
                    if second_category and second_category not in first_category_map:
                        missing_second_categories.add(second_category)
                        if row_idx <= data_start_row + 2:
                            print(f"[DEBUG] 预扫描第{row_idx}行：发现缺失的第二分类 '{second_category}'")
            
            # 如果有缺失的第二分类，提示用户
            if missing_second_categories:
                missing_list = '\n'.join(f'• {cat}' for cat in sorted(missing_second_categories))
                from core.custom_message_box import CustomMessageBox
                reply = CustomMessageBox.warning(
                    self,
                    "发现缺失的分类",
                    f"Excel文件中发现以下第二分类未在分类管理中配置：\n\n{missing_list}\n\n请先在设置→分类管理中添加这些分类，然后重新导入。",
                    ("确定",)
                )
                print(f"[WARNING] 发现缺失的第二分类: {missing_second_categories}")
                return  # 中断导入
            
            # 从数据开始行读取数据
            for row_idx, row in enumerate(sheet.iter_rows(min_row=data_start_row), start=data_start_row):
                if not any(cell.value for cell in row):
                    continue  # 跳过空行
                
                if row_idx <= data_start_row + 2:  # 只显示前几行的调试信息
                    print(f"[DEBUG] 处理第{row_idx}行，数据: {[cell.value for cell in row]}")
                
                achievement = {}
                
                # 1. 绝对编号列：暂时不生成，在最后统一重新生成
                abs_id = ''
                if '绝对编号' in col_index:
                    abs_id_value = row[col_index['绝对编号']].value
                    abs_id = str(abs_id_value).strip() if abs_id_value else ''
                achievement['绝对编号'] = abs_id
                
                # 2. 名称列：去除「隐藏成就」
                name_value = row[col_index['名称']].value
                name = str(name_value).strip() if name_value else ''
                if '「隐藏成就」' in name:
                    name = name.replace('「隐藏成就」', '').strip()
                achievement['名称'] = name
                
                # 3. 描述列
                desc_value = row[col_index['描述']].value if '描述' in col_index else ''
                description = str(desc_value).strip() if desc_value else ''
                achievement['描述'] = description
                
                # 4. 版本列：智能处理小数
                version_value = row[col_index['版本']].value if '版本' in col_index else ''
                version = str(version_value).strip() if version_value else ''
                if version:
                    if '.' in version:
                        pass  # 已经有小数点，保持原样
                    else:
                        version = f"{version}.0"
                achievement['版本'] = version
                
                # 5. 奖励列：纯数字拼接"星声*"
                reward_value = row[col_index['奖励']].value if '奖励' in col_index else ''
                reward = str(reward_value).strip() if reward_value else ''
                if reward.isdigit():
                    reward = f"星声*{reward}"
                achievement['奖励'] = reward
                
                # 6. 是否隐藏列：简化判断，只判断是否包含"隐藏"
                if '是否隐藏' in col_index:
                    hidden_value = row[col_index['是否隐藏']].value
                    is_hidden = str(hidden_value).strip() if hidden_value else ''
                    achievement['是否隐藏'] = '隐藏' if '隐藏' in is_hidden else ''
                else:
                    achievement['是否隐藏'] = ''
                
                # 7. 第二分类列：必须有
                second_category_value = row[col_index['第二分类']].value
                second_category = str(second_category_value).strip() if second_category_value else ''
                if not second_category:
                    raise Exception(f"第{row_idx}行：第二分类不能为空")
                achievement['第二分类'] = second_category
                
                # 8. 第一分类列：如果为空，根据第二分类映射自动补充
                first_category = ''
                has_first_category_column = '第一分类' in col_index
                
                if has_first_category_column:
                    first_category_value = row[col_index['第一分类']].value
                    first_category = str(first_category_value).strip() if first_category_value else ''

                # 如果第一分类为空，根据第二分类映射获取第一分类
                if not first_category:
                    print(f"[DEBUG] 第{row_idx}行：尝试为第二分类 '{second_category}' 查找第一分类")
                    print(f"[DEBUG] 第{row_idx}行：first_category_map包含的分类: {list(first_category_map.keys())}")
                    
                    # 尝试精确匹配
                    first_category = first_category_map.get(second_category, '')
                    
                    # 如果精确匹配失败，尝试去除空格后匹配
                    if not first_category:
                        trimmed_second_category = second_category.strip()
                        first_category = first_category_map.get(trimmed_second_category, '')
                        if first_category:
                            print(f"[DEBUG] 第{row_idx}行：去除空格后匹配成功")
                    
                    if first_category:
                        print(f"[INFO] 第{row_idx}行：自动补充第一分类 '{first_category}'")
                    else:
                        print(f"[WARNING] 第{row_idx}行：无法找到第二分类 '{second_category}' 对应的第一分类")
                
                achievement['第一分类'] = first_category
                
                # 9. 编号列：暂时不生成，在最后统一重新生成
                serial_number = ''
                if '编号' in col_index:
                    serial_value = row[col_index['编号']].value
                    serial_number = str(serial_value).strip() if serial_value else ''
                
                # 编号列：暂时不生成，在最后统一重新生成
                achievement['编号'] = serial_number
                
                # 10. 获取状态列
                if '获取状态' in col_index:
                    status_value = row[col_index['获取状态']].value
                    achievement['获取状态'] = str(status_value).strip() if status_value else ''
                else:
                    achievement['获取状态'] = ''
                
                # 11. 成就组ID列
                if '成就组ID' in col_index:
                    group_id_value = row[col_index['成就组ID']].value
                    achievement['成就组ID'] = str(group_id_value).strip() if group_id_value else ''
                else:
                    achievement['成就组ID'] = ''
                
                # 12. 互斥成就列
                if '互斥成就' in col_index:
                    exclusive_value = row[col_index['互斥成就']].value
                    if exclusive_value:
                        # 如果是字符串，尝试分割
                        if isinstance(exclusive_value, str):
                            # 尝试按逗号、分号或空格分割
                            import re
                            parts = re.split(r'[,;，；\s]+', exclusive_value.strip())
                            achievement['互斥成就'] = [part.strip() for part in parts if part.strip()]
                        else:
                            achievement['互斥成就'] = [str(exclusive_value)]
                    else:
                        achievement['互斥成就'] = []
                else:
                    achievement['互斥成就'] = []
                
                achievements.append(achievement)
            
            workbook.close()
            
            # 分离基础成就和用户进度
            base_achievements = []
            user_progress = {}
            
            # 获取当前用户
            current_user = config.get_current_user()
            
            for achievement in achievements:
                # 创建基础成就副本（不包含获取状态）
                base_achievement = achievement.copy()
                # 只删除获取状态，保留其他所有字段（包括成就组ID和互斥成就）
                if '获取状态' in base_achievement:
                    del base_achievement['获取状态']
                base_achievements.append(base_achievement)
                
                # 如果有获取状态，添加到用户进度
                # 使用编号作为键（用户进度文件使用编号）
                if achievement.get('获取状态'):
                    serial_number = achievement.get('编号', '')
                    if serial_number:
                        user_progress[serial_number] = {
                            '获取状态': achievement['获取状态']
                        }
            
            # 重新生成编号和绝对编号（与爬虫保持一致）
            print("[INFO] 正在重新生成编号和绝对编号...")
            base_achievements = self._smart_reencode_achievements(base_achievements)
            print("[SUCCESS] 编号和绝对编号重新生成完成")
            
            # 将用户进度映射到新的编号
            if user_progress:
                updated_user_progress = {}
                for achievement in base_achievements:
                    serial_number = achievement.get('编号', '')
                    if serial_number and serial_number in user_progress:
                        updated_user_progress[serial_number] = {
                            '获取状态': user_progress[serial_number]['获取状态']
                        }
                    else:
                        # 检查是否是名称+描述的临时键
                        name = achievement.get('名称', '')
                        desc = achievement.get('描述', '')
                        key = f"{name}|{desc}"
                        
                        if key in user_progress:
                            serial_number = achievement.get('编号', '')
                            if serial_number:
                                updated_user_progress[serial_number] = {
                                    'status': user_progress[key]['status'],
                                    'timestamp': user_progress[key]['timestamp']
                                }
                
                user_progress = updated_user_progress
                print(f"[INFO] 用户进度已映射到新的编号，共 {len(user_progress)} 条")
            
            # 保存基础成就数据
            base_file = get_resource_path("resources/base_achievements.json")
            with open(base_file, 'w', encoding='utf-8') as f:
                json.dump(base_achievements, f, ensure_ascii=False, indent=2)
            
            # 保存用户进度数据
            if user_progress:
                config.save_user_progress(current_user, user_progress)
            
            # 更新管理器的数据
            self.manager.achievements = base_achievements
            self.manager.filtered_achievements = base_achievements.copy()
            
            # 调试：检查第一分类是否正确补充
            for i, achievement in enumerate(base_achievements[:3]):
                print(f"[DEBUG] 保存前成就{i+1}: 第一分类='{achievement.get('第一分类', '空')}', 第二分类='{achievement.get('第二分类', '空')}'")
            
            # 将用户进度状态合并到基础成就数据中，用于表格显示
            if user_progress:
                for achievement in base_achievements:
                    serial_number = achievement.get('编号', '')
                    if serial_number and serial_number in user_progress:
                        achievement['获取状态'] = user_progress[serial_number]['获取状态']
                    else:
                        achievement['获取状态'] = '未完成'
            else:
                # 如果没有用户进度，全部设为未完成
                for achievement in base_achievements:
                    achievement['获取状态'] = '未完成'
            
            # 更新表格显示
            self.manager_table.load_data(base_achievements)
            
            # 更新筛选器
            self.update_filters()
            
            # 更新统计
            self.update_statistics()
            
            print(f"[SUCCESS] Excel导入完成，共 {len(base_achievements)} 条基础成就，{len(user_progress)} 条用户进度")
            
            # 检查是否有缺失的分类，如果有则中断导入
            if hasattr(self, 'missing_categories') and self.missing_categories:
                missing_list = sorted(list(self.missing_categories))
                missing_str = "、".join(missing_list)
                
                # 显示错误提示，中断导入
                error_msg = f"发现未配置的第二分类: {missing_str}\n\n"
                error_msg += "请在 设置→分类管理 中将这些分类添加到对应的第一分类下，然后重新导入。"
                
                print(f"[ERROR] 导入中断：发现未配置的分类: {missing_str}")
                print(f"[INFO] 请在 设置→分类管理 中添加这些分类后重新导入")
                
                # 显示错误提示
                from core.custom_message_box import CustomMessageBox
                CustomMessageBox.warning(self, "导入中断", error_msg)
                
                # 不更新数据，保持原状
                return
            
            print(f"[SUCCESS] Excel导入完成，共 {len(base_achievements)} 条基础成就，{len(user_progress)} 条用户进度")
            show_notification(self, f"导入成功，共 {len(base_achievements)} 条成就数据")
            
        except Exception as e:
            print(f"[ERROR] 导入Excel失败: {str(e)}")
            raise Exception(f"导入Excel失败: {str(e)}")
    
    def export_excel(self):
        """导出Excel文件（包含基础成就和用户进度）"""
        # 获取动态文件名
        dynamic_filename = self.get_dynamic_export_filename()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出Excel文件", dynamic_filename, "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                self.export_to_excel(file_path)
            except Exception as e:
                print(f"[ERROR] 导出Excel失败: {str(e)}")
                show_notification(self, f"导出Excel失败: {str(e)}")
    
    def export_to_excel(self, file_path):
        """将基础成就和用户进度合并导出为Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 创建工作簿
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "成就数据"
            
            # 获取基础成就数据
            base_achievements = config.load_base_achievements()
            
            # 获取当前用户进度数据
            current_user = config.get_current_user()
            user_progress = config.load_user_progress(current_user)
            
            # 创建用户进度映射
            progress_map = {item.get('绝对编号', ''): item.get('获取状态', '') 
                          for item in self.manager.achievements if item.get('绝对编号')}
            
            # 合并数据
            merged_data = []
            for achievement in base_achievements:
                merged_item = achievement.copy()
                # 添加获取状态
                abs_id = achievement.get('绝对编号', '')
                merged_item['获取状态'] = progress_map.get(abs_id, '')
                merged_data.append(merged_item)
            
            # 定义列顺序
            column_order = [
                '绝对编号', '版本', '第一分类', '第二分类', '编号', 
                '名称', '描述', '奖励', '是否隐藏', '获取状态',
                '成就组ID', '互斥成就'
            ]
            
            # 在第一行添加导出时间信息，使用与统计信息相同的逻辑
            current_user = config.get_current_user()
            users = config.get_users()
            user_data = users.get(current_user, {})
            uid = user_data.get('uid', current_user) if isinstance(user_data, dict) else current_user
            statistics = self.calculate_statistics(merged_data)
            info_text = f"导出时间: {self.get_current_time()} | 用户: {current_user} | UID: {uid} | 成就总数: {statistics['total']} | 已完成: {statistics['completed']} | 未完成: {statistics['incomplete']} | 隐藏成就: {statistics['hidden']} | 暂不可获取: {statistics['unavailable']} | 多选一成就: {statistics['multi_choice']}"
            info_cell = sheet.cell(row=1, column=1, value=info_text)
            info_cell.font = Font(size=10, color="666666")
            info_cell.alignment = Alignment(horizontal="left", vertical="center")
            # 合并信息单元格
            sheet.merge_cells(start_row=1, start_column=1, 
                            end_row=1, end_column=len(column_order))
            
            # 设置表头（现在在第二行）
            for col, header in enumerate(column_order, 1):
                cell = sheet.cell(row=2, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 填充数据（从第三行开始）
            for row_idx, item in enumerate(merged_data, 3):
                for col_idx, field_name in enumerate(column_order, 1):
                    value = item.get(field_name, '')
                    
                    # 特殊处理互斥成就列 - 保持数组格式，但Excel中显示为逗号分隔的字符串
                    if field_name == '互斥成就' and isinstance(value, list):
                        # 使用逗号分隔，这样导入时可以正确解析
                        value = ', '.join(str(item) for item in value) if value else ''
                    
                    cell = sheet.cell(row=row_idx, column=col_idx, value=str(value))
                    
                    # 设置边框
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                    
                    # 特殊格式化
                    if field_name == '名称':
                        cell.font = Font(bold=True)
                    elif field_name == '是否隐藏' and value == '隐藏':
                        cell.font = Font(color="FF9900")  # 橙色
                    elif field_name == '奖励' and '星声*' in str(value):
                        # 根据星声数量设置不同颜色
                        if '20' in str(value):
                            cell.font = Font(color="FF6B35")  # 橙色
                        elif '10' in str(value):
                            cell.font = Font(color="4ECDC4")  # 青色
                        elif '5' in str(value):
                            cell.font = Font(color="45B7D1")  # 蓝色
                    elif field_name == '获取状态':
                        # 根据获取状态设置颜色
                        if value == '已完成':
                            cell.font = Font(color="00AA00")  # 绿色
                            cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")  # 浅绿色背景
                        elif value == '已占用':
                            cell.font = Font(color="FF6B35")  # 橙红色
                            cell.fill = PatternFill(start_color="FFE8E0", end_color="FFE8E0", fill_type="solid")  # 浅橙色背景
                        elif value == '未完成':
                            cell.font = Font(color="888888")  # 灰色
                            cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # 浅灰色背景
                        elif value == '暂不可获取':
                            cell.font = Font(color="999999")  # 深灰色
                            cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")  # 深灰色背景
            
            # 调整列宽
            column_widths = {
                '绝对编号': 12,
                '版本': 10,
                '第一分类': 15,
                '第二分类': 20,
                '编号': 12,
                '名称': 25,
                '描述': 40,
                '奖励': 15,
                '是否隐藏': 10,
                '获取状态': 10,
                '成就组ID': 15,
                '互斥成就': 20
            }
            
            for col_idx, field_name in enumerate(column_order, 1):
                sheet.column_dimensions[get_column_letter(col_idx)].width = column_widths.get(field_name, 15)
            
            # 时间信息已经移到顶部第一行，无需在底部重复添加
            
            # 添加下拉框和条件格式
            self.add_excel_validation_and_formatting(sheet, len(merged_data), column_order)
            
            # 保存文件
            workbook.save(file_path)
            print(f"[SUCCESS] Excel数据已导出到: {file_path}")
            print(f"[INFO] 包含 {len(merged_data)} 条成就数据")
            show_notification(self, f"导出成功，共 {len(merged_data)} 条成就数据")
            
        except Exception as e:
            print(f"[ERROR] 导出Excel失败: {str(e)}")
            raise Exception(f"导出Excel失败: {str(e)}")
    
    def get_dynamic_export_filename(self):
        """生成动态导出文件名"""
        try:
            # 获取基础成就数据
            base_achievements = config.load_base_achievements()
            
            # 提取所有版本号并排序
            versions = set()
            for achievement in base_achievements:
                version = achievement.get('版本', '')
                if version:
                    versions.add(version)
            
            if versions:
                # 排序版本号
                sorted_versions = sorted(versions, key=lambda x: float(x))
                min_version = sorted_versions[0]
                max_version = sorted_versions[-1]
                
                if min_version == max_version:
                    version_range = f"v{min_version}"
                else:
                    version_range = f"v{min_version}-{max_version}"
            else:
                version_range = "全版本"
            
            # 获取当前用户信息
            current_user = config.get_current_user()
            users = config.get_users()
            user_data = users.get(current_user, {})
            uid = user_data.get('uid', current_user) if isinstance(user_data, dict) else current_user
            
            # 生成文件名
            filename = f"鸣潮{version_range}全成就数据_uid{uid}.xlsx"
            return filename
            
        except Exception as e:
            print(f"[ERROR] 生成动态文件名失败: {str(e)}")
            return "成就数据.xlsx"

    def add_excel_validation_and_formatting(self, sheet, data_rows, column_order):
        """为Excel添加下拉框和条件格式"""
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.utils import get_column_letter
            
            # 移除分类配置加载，因为不再需要下拉框
            
            # 定义下拉框选项
            status_options = '已完成,已占用,未完成,暂不可获取'
            reward_options = '星声*5,星声*10,星声*20'
            
            # 为每一列添加下拉框
            for col_idx, field_name in enumerate(column_order, 1):
                col_letter = get_column_letter(col_idx)
                
                # 获取状态列 - 添加下拉框
                if field_name == '获取状态':
                    dv = DataValidation(type="list", formula1=f'"{status_options}"', allow_blank=True)
                    dv.error = '请从下拉列表中选择有效的状态'
                    dv.errorTitle = '输入错误'
                    # 应用到数据区域（从第3行到最后一行）
                    sheet.add_data_validation(dv)
                    dv.add(f"{col_letter}3:{col_letter}{data_rows + 2}")
                
                # 第一分类和第二分类列 - 移除下拉框，允许自由输入
            
            # 移除条件格式功能，因为Excel动态条件格式支持有限
            
            # 为奖励列添加下拉框
            reward_col_idx = column_order.index('奖励') + 1 if '奖励' in column_order else 0
            if reward_col_idx > 0:
                reward_col_letter = get_column_letter(reward_col_idx)
                # 添加下拉框验证
                dv = DataValidation(type="list", formula1=f'"{reward_options}"', allow_blank=True)
                dv.error = '请从下拉列表中选择有效的奖励数量'
                dv.errorTitle = '输入错误'
                dv.prompt = '请选择奖励数量'
                dv.promptTitle = '提示'
                sheet.add_data_validation(dv)
                dv.add(f"{reward_col_letter}3:{reward_col_letter}{data_rows + 2}")
            
            print(f"[INFO] 已添加Excel下拉框 - 数据行数: {data_rows}")
            print(f"[DEBUG] 奖励列索引: {reward_col_idx}")
            
        except Exception as e:
            print(f"[WARNING] 添加Excel验证和格式时出错: {str(e)}")

    def calculate_statistics(self, data):
        """计算统计信息，与update_statistics逻辑相同"""
        # 正确统计总计（考虑成就组）
        total_groups = set()
        total_achievements = 0
        for achievement in data:
            group_id = achievement.get('成就组ID')
            if group_id:
                total_groups.add(group_id)
            else:
                total_achievements += 1
        total = len(total_groups) + total_achievements
        
        # 统计每个成就组的状态
        group_status = {}  # group_id -> {'status': 'completed'/'incomplete'/'unavailable', 'has_hidden': bool}
        for achievement in data:
            status = achievement.get('获取状态', '') or '未完成'
            group_id = achievement.get('成就组ID')
            is_hidden = achievement.get('是否隐藏') == '隐藏'
            
            if group_id:
                if group_id not in group_status:
                    group_status[group_id] = {'status': status, 'has_hidden': is_hidden}
                else:
                    # 更新状态：已完成 > 暂不可获取 > 未完成
                    current = group_status[group_id]['status']
                    if status == '已完成' or (status == '暂不可获取' and current != '已完成'):
                        group_status[group_id]['status'] = status
                    if is_hidden:
                        group_status[group_id]['has_hidden'] = True
        
        # 统计已完成（考虑成就组）
        completed = 0
        for group_id, info in group_status.items():
            if info['status'] == '已完成':
                completed += 1
        
        # 统计普通已完成成就
        for achievement in data:
            if achievement.get('获取状态', '') == '已完成' and not achievement.get('成就组ID'):
                completed += 1
        
        # 统计未完成（考虑成就组）
        incomplete = 0
        for group_id, info in group_status.items():
            if info['status'] == '未完成':
                incomplete += 1
        
        # 统计普通未完成成就
        for achievement in data:
            status = achievement.get('获取状态', '') or '未完成'
            if status == '未完成' and not achievement.get('成就组ID'):
                incomplete += 1
        
        # 统计隐藏成就（考虑成就组）
        hidden = 0
        processed_hidden_groups = set()
        for achievement in data:
            is_hidden = achievement.get('是否隐藏') == '隐藏'
            group_id = achievement.get('成就组ID')
            
            if is_hidden:
                if group_id:
                    # 成就组：只计算一次
                    if group_id not in processed_hidden_groups:
                        hidden += 1
                        processed_hidden_groups.add(group_id)
                else:
                    # 普通成就
                    hidden += 1
        
        # 统计暂不可获取数量（考虑成就组）
        unavailable = 0
        processed_unavailable_groups = set()
        for achievement in data:
            status = achievement.get('获取状态', '')
            group_id = achievement.get('成就组ID')
            
            if status == '暂不可获取':
                if group_id:
                    # 成就组：只计算一次
                    if group_id not in processed_unavailable_groups:
                        unavailable += 1
                        processed_unavailable_groups.add(group_id)
                else:
                    # 普通成就
                    unavailable += 1
        
        # 统计多选一数量（每个组只计算一次）
        multi_choice_groups = set()
        for achievement in data:
            group_id = achievement.get('成就组ID')
            if group_id:
                multi_choice_groups.add(group_id)
        multi_choice = len(multi_choice_groups)
        
        return {
            'total': total,
            'completed': completed,
            'incomplete': incomplete,
            'hidden': hidden,
            'unavailable': unavailable,
            'multi_choice': multi_choice
        }

    # Excel动态下拉框实现复杂，暂时使用静态下拉框，用户需要手动确保分类匹配

    def get_current_time(self):
        """获取当前时间字符串"""
        from datetime import datetime
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def import_json(self):
            """导入JSON文件"""
            # 添加确认对话框
            from core.custom_message_box import CustomMessageBox
            reply = CustomMessageBox.question(
                self, 
                "确认导入", 
                "确定要导入JSON文件吗？\n此操作将覆盖当前所有成就数据和用户进度，且不可撤销！\n\n建议先备份当前数据。",
                ("确定", "取消")
            )
            
            if reply != CustomMessageBox.Yes:
                print("[INFO] 用户取消了JSON导入操作")
                return
            
            file_path, _ = QFileDialog.getOpenFileName(
                self, "导入JSON文件", "", "JSON Files (*.json)"
            )
            if file_path:
                try:
                    print(f"[INFO] 开始导入 JSON 文件: {file_path}")

                    with open(file_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)

                    achievements = []
                    needs_reencoding = False  # 是否需要重新编码

                    if isinstance(data, list):
                        # 数组格式：直接是成就列表
                        for achievement in data:
                            # 检查是否需要重新编码
                            if self._needs_reencoding(achievement):
                                needs_reencoding = True
                            # 标准化字段名
                            standardized = _standardize_achievement_fields(achievement)
                            achievements.append(standardized)
                    elif isinstance(data, dict):
                        # 字典格式：可能是进度数据或完整数据
                        if len(list(data.keys())[0]) <= 10:  # 可能是进度数据
                            # 进度数据格式，需要与基础数据合并
                            base_achievements = config.load_base_achievements()
                            for base_achievement in base_achievements:
                                achievement = base_achievement.copy()
                                achievement_id = achievement.get("编号", "")

                                # 添加用户进度
                                if achievement_id in data:
                                    progress = data[achievement_id]
                                    achievement["获取状态"] = progress.get("获取状态", "")
                                else:
                                    achievement["获取状态"] = ""

                                # 转换为内部使用的字段名
                                if "是否隐藏" in achievement:
                                    achievement["is_hidden"] = achievement["是否隐藏"] == "隐藏"

                                achievements.append(achievement)
                        else:
                            # 完整数据格式，直接使用
                            for achievement_id, achievement_data in data.items():
                                achievement = achievement_data.copy()
                                achievement["编号"] = achievement_id

                                # 检查是否需要重新编码
                                if self._needs_reencoding(achievement):
                                    needs_reencoding = True

                                # 标准化字段名
                                standardized = _standardize_achievement_fields(achievement)
                                achievements.append(standardized)

                    if achievements:
                        # 如果需要重新编码，进行智能重新排序和编码
                        if needs_reencoding:
                            print("[INFO] 检测到需要重新编码的数据，正在优化排序和编码...")
                            achievements = self._smart_reencode_achievements(achievements)

                        # 更新管理器数据
                        self.manager.load_data(achievements)

                        # 更新表格
                        self.manager_table.load_data(achievements)

                        # 更新筛选器
                        self.update_filters()

                        # 更新统计
                        self.update_statistics()

                        # 保存为JSON
                        self.save_to_json()

                        status_msg = f"成功导入 {len(achievements)} 条成就数据"
                        if needs_reencoding:
                            status_msg += "（已优化排序和编码）"
                        print(f"[SUCCESS] {status_msg}")

                        # 显示提示
                        show_notification(self, status_msg)
                    else:
                        print("[WARNING] 导入的文件中没有数据")

                except Exception as e:
                    print(f"[ERROR] 导入失败: {str(e)}")
                    import traceback
                    traceback.print_exc()
    
    def _needs_reencoding(self, achievement):
        """判断成就是否需要重新编码"""
        # 通过获取状态字段判断是否为历史数据
        status = achievement.get('获取状态', '')
        
        # 如果获取状态是旧格式（空字符串、completed、unavailable），需要重新编码
        if status in ['', 'completed', 'unavailable']:
            return True
        
        return False
    
    def _smart_reencode_achievements(self, achievements):
        """智能重新编码成就，优化排序"""
        # 按照要求的排序规则排序：
        # 1. 保持相对稳定（现有编号尽可能保持不变）
        # 2. 第一分类排序
        # 3. 第二分类排序  
        # 4. 版本号按浮点型正序
        
        # 获取分类配置
        from core.config import config
        try:
            category_config = config.load_category_config()
            if not isinstance(category_config, dict):
                category_config = {}
            first_categories = category_config.get("first_categories", {})
            second_categories = category_config.get("second_categories", {})
        except Exception as e:
            print(f"[ERROR] 加载分类配置失败: {str(e)}")
            first_categories = {}
            second_categories = {}
        
        def get_sort_key(achievement):
            """获取排序键"""
            # 第一分类排序
            first_cat = achievement.get('第一分类', '')
            first_order = first_categories.get(first_cat, 999)
            
            # 第二分类排序
            second_cat = achievement.get('第二分类', '')
            first_cat_second = second_categories.get(first_cat, {})
            second_order = int(first_cat_second.get(second_cat, 999)) if second_cat in first_cat_second else 999
            
            # 版本号排序（浮点型正序）
            version_str = achievement.get('版本', '0.0')
            try:
                version = float(version_str)
            except ValueError:
                version = 0.0
            
            # 原编号（用于保持相对稳定）
            original_id = achievement.get('编号', '99999999')
            
            return (first_order, second_order, version, original_id)
        
        # 创建第二分类到第一分类的映射
        first_category_map = {}
        for first_cat, second_cats in second_categories.items():
            for second_cat in second_cats:
                first_category_map[second_cat] = first_cat
        
        # 按新规则排序
        sorted_achievements = sorted(achievements, key=get_sort_key)
        
        # 重新分配编号
        current_numbers = {}
        for achievement in sorted_achievements:
            first_cat = achievement.get('第一分类', '')
            second_cat = achievement.get('第二分类', '')
            
            # 如果第一分类为空，根据第二分类映射自动补充
            if not first_cat and second_cat:
                first_cat = first_category_map.get(second_cat, '')
                if first_cat:
                    achievement['第一分类'] = first_cat
                    print(f"[INFO] 自动补充第一分类 '{first_cat}' 用于第二分类 '{second_cat}'")
            
            if not first_cat or not second_cat:
                achievement['编号'] = ''
                continue
            
            # 获取第一分类排序号
            first_category_order = first_categories.get(first_cat, 1)
            
            # 获取第二分类后缀
            suffix = second_categories.get(first_cat, {}).get(second_cat, '10')
            
            # 生成完整前缀：第一分类(1位) + 第二分类后缀(补齐到3位)
            suffix_padded = f"{int(suffix):03d}"
            full_prefix = f"{first_category_order}{suffix_padded}"
            
            # 获取当前序号
            category_key = (first_cat, second_cat)
            current_num = current_numbers.get(category_key, 1)
            
            # 生成编号：4位分类码 + 4位序号
            achievement['编号'] = f"{full_prefix}{current_num:04d}"
            
            # 更新序号
            current_numbers[category_key] = current_num + 1
        
        # 重新生成绝对编号（按最终排序顺序从1开始递增）
        for index, achievement in enumerate(sorted_achievements, start=1):
            achievement['绝对编号'] = str(index)
        
        return sorted_achievements
    
    def process_full_field_data(self, data):
        """处理全字段数据，兼容新旧获取状态格式"""
        achievements = []
        
        # 获取状态映射
        status_mapping = {
            'completed': '已完成',
            '': '未完成',
            'unavailable': '暂不可获取'
        }
        
        for item in data:
            achievement = {}
            
            # 处理字段映射
            field_mapping = {
                '绝对编号': '绝对编号',
                'version': '版本',
                '第一分类': '第一分类',
                '第二分类': '第二分类',
                '编号': '编号',
                'name': '名称',
                'description': '描述',
                'reward': '奖励',
                '是否隐藏': '是否隐藏',
                '获取状态': '获取状态'
            }
            
            for key, value in item.items():
                mapped_key = field_mapping.get(key, key)
                achievement[mapped_key] = value
            
            # 处理是否隐藏字段
            if 'is_hidden' in achievement:
                achievement['是否隐藏'] = '隐藏' if achievement['is_hidden'] else ''
            elif '是否隐藏' not in achievement:
                achievement['是否隐藏'] = ''
            
            # 兼容新旧获取状态格式
            if '获取状态' in achievement:
                old_status = achievement['获取状态']
                # 如果是旧格式，进行转换
                if old_status in status_mapping:
                    achievement['获取状态'] = status_mapping[old_status]
                # 如果已经是新格式，保持不变
            
            # 兼容版本字段格式
            if '版本' in achievement:
                version = achievement['版本']
                # 如果是整数形式的字符串，转换为浮点数格式
                try:
                    if version and '.' not in version:
                        # 检查是否为纯数字
                        if version.isdigit():
                            achievement['版本'] = f"{int(version)}.0"
                        else:
                            # 尝试转换为浮点数
                            float_val = float(version)
                            achievement['版本'] = f"{float_val}.0" if float_val == int(float_val) else str(float_val)
                except (ValueError, TypeError):
                    # 如果转换失败，保持原值
                    pass
            
            achievements.append(achievement)
        
        return achievements
    
    def export_full_json(self):
        """导出全字段 JSON 文件"""
        if not self.manager.filtered_achievements:
            print("[WARNING] 没有数据可导出")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出全字段数据", "wuthering_waves_achievements_full.json", "JSON Files (*.json)"
        )
        
        if file_path:
            try:
                print(f"[INFO] 开始导出全字段数据: {file_path}")
                
                # 准备全字段数据
                export_data = []
                for achievement in self.manager.filtered_achievements:
                    export_data.append({
                        '绝对编号': achievement.get('绝对编号', ''),
                        '版本': achievement.get('版本', ''),
                        '第一分类': achievement.get('第一分类', ''),
                        '第二分类': achievement.get('第二分类', ''),
                        '编号': achievement.get('编号', ''),
                        '名称': achievement.get('名称', ''),
                        '描述': achievement.get('描述', ''),
                        '奖励': achievement.get('奖励', ''),
                        '是否隐藏': achievement.get('是否隐藏', ''),
                        '获取状态': achievement.get('获取状态', '')
                    })
                
                # 保存到 JSON
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(export_data, f, ensure_ascii=False, indent=2)
                
                file_size = os.path.getsize(file_path)
                print(f"[SUCCESS] 全字段数据已导出到: {file_path}")
                print(f"[SUCCESS] 文件大小: {file_size} 字节")
                print(f"[INFO] 包含 {len(export_data)} 条成就数据")
                
            except Exception as e:
                print(f"[ERROR] 导出失败: {str(e)}")
                import traceback
                traceback.print_exc()
    
    def update_filters(self):
        """更新筛选器选项"""
        versions = set()
        first_categories = set()
        second_categories = set()
        
        for achievement in self.manager.achievements:
            versions.add(achievement.get('版本', ''))
            first_categories.add(achievement.get('第一分类', ''))
            second_categories.add(achievement.get('第二分类', ''))
        
        # 更新版本下拉框
        self.version_filter.clear()
        self.version_filter.addItem("所有版本")
        
        # 将版本号转换为浮点数进行倒序排序
        valid_versions = []
        for v in versions:
            if v:
                try:
                    # 尝试转换为浮点数进行排序
                    float_val = float(v)
                    valid_versions.append((float_val, v))
                except (ValueError, TypeError):
                    # 如果转换失败，使用原始字符串排序
                    valid_versions.append((-1, v))
        
        # 按浮点数值倒序排序，无法转换的放在最后
        sorted_versions = [v[1] for v in sorted(valid_versions, key=lambda x: x[0], reverse=True)]
        
        for version in sorted_versions:
            self.version_filter.addItem(version)
        
# 获取分类配置
        try:
            category_config = config.load_category_config()
            if not isinstance(category_config, dict):
                category_config = {}
        except Exception as e:
            print(f"[ERROR] 加载分类配置失败: {str(e)}")
            # 尝试直接访问属性
            try:
                if hasattr(config, 'category_config'):
                    category_config = getattr(config, 'category_config', {})
                else:
                    category_config = {}
            except:
                category_config = {}
            print("[INFO] 使用备用配置加载方式")
        first_category_order = category_config.get("first_categories", {})
        
        # 更新第一分类下拉框（按配置顺序）
        self.first_category_filter.clear()
        self.first_category_filter.addItem("全部")
        # 按照配置中的排序顺序添加
        ordered_first_categories = sorted(first_categories, key=lambda x: first_category_order.get(x, 999))
        for category in ordered_first_categories:
            if category:
                self.first_category_filter.addItem(category)
        
        # 更新第二分类下拉框
        self.second_category_filter.clear()
        self.second_category_filter.addItem("全部")
        # 第二分类需要根据当前选中的第一分类来排序
        current_first = self.first_category_filter.currentText()
        if current_first != "全部" and current_first in category_config.get("second_categories", {}):
            second_category_order = category_config["second_categories"].get(current_first, {})
            ordered_second_categories = sorted(second_categories, key=lambda x: int(second_category_order.get(x, 999)))
            for category in ordered_second_categories:
                if category:
                    self.second_category_filter.addItem(category)
        else:
            # 如果没有选中第一分类或选中"全部"，则按字母顺序
            for category in sorted(second_categories):
                if category:
                    self.second_category_filter.addItem(category)
    
    def save_to_json(self):
        """分离保存基础数据和用户进度"""
        try:
            current_user = config.get_current_user()
            users = config.get_users()
            current_user_data = users.get(current_user, {})
            uid = current_user_data.get('uid', current_user) if isinstance(current_user_data, dict) else current_user
            
            # 保存基础成就数据（恢复原始名称）
            achievements_to_save = []
            for achievement in self.manager.achievements:
                # 创建副本，避免修改原始数据
                achievement_copy = achievement.copy()
                # 如果有原始名称，恢复它
                if '原始名称' in achievement_copy:
                    achievement_copy['名称'] = achievement_copy['原始名称']
                    # 删除临时字段
                    del achievement_copy['原始名称']
                achievements_to_save.append(achievement_copy)
            
            if config.save_base_achievements(achievements_to_save):
                print("[SUCCESS] 基础成就数据已保存")
            
            # 准备用户进度数据
            progress_data = {}
            for achievement in self.manager.achievements:
                achievement_id = achievement.get("编号", "")
                status = achievement.get("获取状态", "未完成")
                # 如果状态为空字符串，设置为"未完成"
                if not status:
                    status = "未完成"
                progress_data[achievement_id] = {
                    "获取状态": status
                }
            
            # 保存用户进度数据
            if config.save_user_progress(current_user, progress_data):
                print(f"[SUCCESS] 用户 {current_user} (UID: {uid}) 的进度数据已保存")
            
        except Exception as e:
            print(f"[ERROR] 保存数据失败: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def load_local_data(self):
            """加载本地数据：合并基础数据和用户进度"""
            try:
                # 加载基础成就数据
                base_achievements = config.load_base_achievements()
                if not base_achievements:
                    print("[WARNING] 基础成就数据文件不存在")
                    return
                
                # 加载当前用户进度数据
                current_user = config.get_current_user()
                users = config.get_users()
                current_user_data = users.get(current_user, {})
                uid = current_user_data.get('uid', current_user) if isinstance(current_user_data, dict) else current_user
                
                print(f"[INFO] 加载用户 {current_user} (UID: {uid}) 的进度数据")
                user_progress = config.load_user_progress(current_user)
                
                # 合并数据
                achievements = []
                for base_achievement in base_achievements:
                    achievement = base_achievement.copy()
                    achievement_id = achievement.get("编号", "")
                    
                    # 添加用户进度
                    if achievement_id in user_progress:
                        progress = user_progress[achievement_id]
                        achievement["获取状态"] = progress.get("获取状态", "")
                    else:
                        achievement["获取状态"] = ""
                    
                    # 转换为内部使用的字段名
                    if "是否隐藏" in achievement:
                        achievement["is_hidden"] = achievement["是否隐藏"] == "隐藏"
                    
                    achievements.append(achievement)
                
                print(f"[INFO] 加载了 {len(achievements)} 条成就数据（基础数据 + 用户进度）")
                
                if achievements:
                    self.manager.load_data(achievements)
                    
                    # 更新筛选器
                    self.update_filters()
                    
                    # 更新统计
                    self.update_statistics()
                
            except Exception as e:
                        print(f"[ERROR] 加载本地数据失败: {str(e)}")
                        import traceback
                        traceback.print_exc()
                
    def apply_theme(self, theme):
        """应用主题"""
        # 更新按钮样式
        if hasattr(self, 'import_btn'):
            self.import_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'import_excel_btn'):
            self.import_excel_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'export_excel_btn'):
            self.export_excel_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'export_json_btn'):
            self.export_json_btn.setStyleSheet(get_button_style(theme))
        
        # 更新文字样式
        if hasattr(self, 'total_label'):
            self.total_label.setStyleSheet(get_font_gray_style(theme))
        if hasattr(self, 'completed_label'):
            self.completed_label.setStyleSheet(get_font_gray_style(theme))
        if hasattr(self, 'incomplete_label'):
            self.incomplete_label.setStyleSheet(get_font_gray_style(theme))
        if hasattr(self, 'hidden_label'):
            self.hidden_label.setStyleSheet(get_font_gray_style(theme))
        
        # 更新输入框样式
        if hasattr(self, 'search_input'):
            from core.styles import BaseStyles
            input_style = BaseStyles.get_text_input_style(theme)
            self.search_input.setStyleSheet(input_style)
        
        # 更新下拉框样式
        if hasattr(self, 'version_filter'):
            from core.styles import BaseStyles
            combo_style = BaseStyles.get_combobox_style(theme)
            self.version_filter.setStyleSheet(combo_style)
        if hasattr(self, 'first_category_filter'):
            from core.styles import BaseStyles
            combo_style = BaseStyles.get_combobox_style(theme)
            self.first_category_filter.setStyleSheet(combo_style)
        if hasattr(self, 'second_category_filter'):
            from core.styles import BaseStyles
            combo_style = BaseStyles.get_combobox_style(theme)
            self.second_category_filter.setStyleSheet(combo_style)
        if hasattr(self, 'priority_filter'):
            from core.styles import BaseStyles
            combo_style = BaseStyles.get_combobox_style(theme)
            self.priority_filter.setStyleSheet(combo_style)
        if hasattr(self, 'hidden_filter'):
            from core.styles import BaseStyles
            combo_style = BaseStyles.get_combobox_style(theme)
            self.hidden_filter.setStyleSheet(combo_style)
        if hasattr(self, 'obtainable_filter'):
            from core.styles import BaseStyles
            combo_style = BaseStyles.get_combobox_style(theme)
            self.obtainable_filter.setStyleSheet(combo_style)
        
        # 更新表格样式
        if hasattr(self, 'manager_table'):
            from core.styles import BaseStyles, get_scrollbar_style
            table_style = BaseStyles.get_text_input_style(theme)
            scrollbar = get_scrollbar_style(theme)
            self.manager_table.setStyleSheet(table_style + scrollbar)


def show_notification(parent, message):
    """显示右上角自动关闭的提示"""
    from PySide6.QtWidgets import QLabel
    from PySide6.QtCore import QTimer, Qt
    from core.config import config
    from core.styles import get_notification_style
    
    # 创建提示标签
    notification = QLabel(message)
    notification.setStyleSheet(get_notification_style(config.theme))
    
    # 获取主窗口
    main_window = None
    try:
        from core.main_window import TemplateMainWindow
        for widget in QApplication.topLevelWidgets():
            if isinstance(widget, TemplateMainWindow):
                main_window = widget
                break
    except:
        pass
    
    if not main_window:
        print("[WARNING] 无法找到主窗口，提示将显示在当前窗口")
        parent = parent
    else:
        parent = main_window
    
    # 设置提示为独立窗口，避免被父窗口影响
    notification.setWindowFlags(Qt.WindowType.ToolTip | Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint)
    notification.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose)
    notification.adjustSize()
    
    # 计算右上角位置（使用父窗口的全局坐标）
    parent_rect = parent.geometry()
    notification_width = 300  # 固定宽度
    notification_height = 60  # 固定高度
    
    x = parent_rect.x() + parent_rect.width() - notification_width - 20
    y = parent_rect.y() + 60  # 状态栏下方
    
    notification.setGeometry(x, y, notification_width, notification_height)
    notification.setWordWrap(True)
    notification.setAlignment(Qt.AlignCenter)
    notification.show()
    
    # 创建淡出动画，保持引用避免被垃圾回收
    fade_timer = QTimer()
    fade_timer.setSingleShot(True)
    fade_timer.timeout.connect(lambda: fade_out_notification(notification, parent))
    fade_timer.start(3000)  # 3秒后开始淡out
    
    # 存储动画引用
    if not hasattr(parent, 'active_notification_timers'):
        parent.active_notification_timers = []
    parent.active_notification_timers.append(fade_timer)
    
    # 存储引用以避免被垃圾回收
    if not hasattr(parent, 'active_notifications'):
        parent.active_notifications = []
    parent.active_notifications.append(notification)

def fade_out_notification(notification, parent):
    """淡出提示"""
    from PySide6.QtCore import QPropertyAnimation, QEasingCurve
    
    # 检查通知是否还存在
    try:
        if notification is None or not notification.isVisible():
            return
    except RuntimeError:
        # 对象已被删除
        return
    
    # 创建透明度动画
    fade_animation = QPropertyAnimation(notification, b"windowOpacity")
    fade_animation.setDuration(500)  # 0.5秒淡出
    fade_animation.setStartValue(1.0)
    fade_animation.setEndValue(0.0)
    fade_animation.setEasingCurve(QEasingCurve.OutQuad)
    
    # 动画完成后删除提示
    def cleanup_notification():
        try:
            if notification and hasattr(notification, 'close'):
                notification.close()
            if notification and hasattr(notification, 'deleteLater'):
                notification.deleteLater()
        except RuntimeError:
            pass  # 对象可能已被删除
        
        # 从活动通知列表中移除
        if hasattr(parent, 'active_notifications') and notification in parent.active_notifications:
            parent.active_notifications.remove(notification)
    
    fade_animation.finished.connect(cleanup_notification)
    fade_animation.start()
    
    # 存储动画引用
    if not hasattr(parent, 'active_notification_animations'):
        parent.active_notification_animations = []
    parent.active_notification_animations.append(fade_animation)