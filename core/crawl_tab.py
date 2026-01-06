from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
                               QTableWidget, QTableWidgetItem, QLineEdit,
                               QGroupBox, QFileDialog)
from PySide6.QtCore import Qt, QThread, Signal, QObject
from PySide6.QtWidgets import QApplication
from PySide6.QtGui import QColor
import json

import requests
from bs4 import BeautifulSoup
import re
import html
import os

from core.config import config
from core.manage_tab import show_notification
from core.signal_bus import signal_bus
from core.styles import (get_button_style, get_font_gray_style)


class AchievementCrawler(QObject):
    """成就爬虫类"""
    progress = Signal(str)
    finished = Signal(list)
    error = Signal(str)

    def __init__(self, target_version=None):
        super().__init__()
        self.target_version = target_version
        self.devcode = ""
        self.token = ""
        
        # 从配置文件加载分类配置
        self.category_config = config.load_category_config()
        self.first_categories = self.category_config.get("first_categories", {})
        self.second_categories = self.category_config.get("second_categories", {})
        
        # 创建第二分类到第一分类的映射
        self.first_category_map = {}
        for first_cat, second_cats in self.second_categories.items():
            for second_cat in second_cats:
                self.first_category_map[second_cat] = first_cat
    
    def _load_auth_config(self):
        """从配置中加载认证信息"""
        self.devcode, self.token = config.get_auth_data()

    def crawl(self):
        try:
            # 加载认证信息
            self._load_auth_config()
            self.progress.emit("正在获取成就数据...")
            data = self.get_achievement_data()
            if data:
                self.progress.emit("解析成就数据...")
                achievements = self.parse_achievements_data(data, self.target_version)
                self.finished.emit(achievements)
            else:
                self.error.emit("获取数据失败")
        except Exception as e:
            self.error.emit(str(e))

    def get_achievement_data(self):
        from core.config import get_resource_path
        import json
        
        cache_file = get_resource_path("resources") / "achievement_cache.json"
        
        if cache_file.exists():
            try:
                with open(cache_file, 'r', encoding='utf-8') as f:
                    cached_data = json.load(f)
                self.progress.emit("使用本地缓存数据...")
                print("[INFO] 使用本地缓存数据")
                return cached_data
            except Exception as e:
                print(f"[WARNING] 读取缓存失败: {str(e)}，将重新请求")
        
        url = "https://api.kurobbs.com/wiki/core/catalogue/item/getEntryDetail"
        headers = {
            'Accept': 'application/json, text/plain, */*',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
            'Connection': 'keep-alive',
            'Content-Type': 'application/x-www-form-urlencoded;charset=UTF-8',
            'Origin': 'https://wiki.kurobbs.com',
            'Referer': 'https://wiki.kurobbs.com/',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'devcode': self.devcode,
            'token': self.token,
            'wiki_type': '9'
        }
        data = {'id': '1220879855033786368'}

        try:
            response = requests.post(url, headers=headers, data=data, timeout=30)
            response.encoding = 'utf-8'
            response_data = response.json()
            
            with open(cache_file, 'w', encoding='utf-8') as f:
                json.dump(response_data, f, ensure_ascii=False, indent=2)
            print(f"[INFO] 已保存缓存到: {cache_file}")
            
            return response_data
        except Exception as e:
            raise Exception(f"网络请求失败: {str(e)}")

    def clean_text(self, text):
        if text is None:
            return ""
        text = re.sub(r'\s+', ' ', text).strip()
        text = html.unescape(text)
        return text

    def parse_html_table(self, html_content):
        achievements = []
        soup = BeautifulSoup(html_content, 'html.parser')
        rows = soup.find_all('tr')

        for row in rows[1:]:
            cells = row.find_all('td')
            if len(cells) >= 5:
                name_text = self.clean_text(cells[0].get_text(strip=True))
                is_hidden = '隐藏成就' in name_text

                if '「隐藏成就」' in name_text:
                    name_text = name_text.replace('「隐藏成就」', '').strip()

                # 从data-filter-tag属性中获取分类信息
                filter_tag = row.get('data-filter-tag', '')
                categories = []
                if filter_tag:
                    # 分割filter_tag获取分类信息
                    tags = filter_tag.split(',')
                    for tag in tags:
                        if tag.startswith('合集-'):
                            category = tag[3:]  # 去掉'合集-'前缀
                            categories.append(category)
                
                # 使用合集作为第二分类
                second_category = categories[0] if categories else self.clean_text(cells[2].get_text(strip=True))
                
                achievement = {
                    '名称': name_text,
                    '版本': self.clean_text(cells[1].get_text(strip=True)),
                    '第一分类': '',  # 第一分类需要后续推断
                    '第二分类': second_category,
                    '描述': self.clean_text(cells[3].get_text(strip=True)),
                    '奖励': self.clean_text(cells[4].get_text(strip=True)),
                    '是否隐藏': '隐藏' if is_hidden else ''
                }
                achievements.append(achievement)
        return achievements
    
    def parse_html_table_with_categories(self, html_content):
                """解析包含折叠分类结构的HTML表格"""
                achievements = []
                soup = BeautifulSoup(html_content, 'html.parser')
                
                # 查找所有的details标签
                details_list = soup.find_all('details', class_='kr-collapse-details')
                
                for details in details_list:
                    # 获取第一分类名称
                    summary = details.find('summary', class_='kr-collapse-summary')
                    if summary:
                        first_category = self.clean_text(summary.get_text(strip=True))
                    else:
                        first_category = ''
                    
                    # 在details内查找表格
                    table = details.find('table', class_='kr-table-filter')
                    if not table:
                        continue
                        
                    # 查找表格中的所有行
                    rows = table.find_all('tr')
                    
                    # 跳过表头行
                    for row in rows[1:]:
                        cells = row.find_all('td')
                        if len(cells) >= 5:
                            name_text = self.clean_text(cells[0].get_text(strip=True))
                            is_hidden = '隐藏成就' in name_text
        
                            if '「隐藏成就」' in name_text:
                                name_text = name_text.replace('「隐藏成就」', '').strip()
        
                            # 从data-filter-tag属性中获取分类信息
                            filter_tag = row.get('data-filter-tag', '')
                            second_category = ''
                            if filter_tag:
                                # 分割filter_tag获取分类信息
                                tags = filter_tag.split(',')
                                for tag in tags:
                                    if tag.startswith('合集-'):
                                        second_category = tag[3:]  # 去掉'合集-'前缀
                                        break
                            
                            # 如果没有从filter-tag获取到，使用第三列
                            if not second_category and len(cells) > 2:
                                second_category = self.clean_text(cells[2].get_text(strip=True))
        
                            achievement = {
                                '名称': name_text,
                                '版本': self.clean_text(cells[1].get_text(strip=True)),
                                '第一分类': first_category,
                                '第二分类': second_category,
                                '描述': self.clean_text(cells[3].get_text(strip=True)),
                                '奖励': self.clean_text(cells[4].get_text(strip=True)),
                                '是否隐藏': '隐藏' if is_hidden else ''
                            }
                            achievements.append(achievement)
                
                return achievements
    
    def parse_achievements_data(self, api_data, target_version=None):
        achievements = []
        try:
            content = api_data.get('data', {}).get('content', {})
            modules = content.get('modules', [])

            for module in modules:
                components = module.get('components', [])
                for component in components:
                    if component.get('type') == 'filter-component':
                        html_content = component.get('content', '')
                        parsed = self.parse_html_table_with_categories(html_content)
                        achievements.extend(parsed)
            
            print(f"[DEBUG] 过滤后剩余 {len(achievements)} 条成就数据")

            # 必须有target_version才进行筛选
            if not target_version:
                raise Exception("必须指定版本号才能爬取数据")
                
            version_filtered = [ach for ach in achievements if ach.get('版本') == target_version]
            print(f"[DEBUG] 版本 {target_version} 筛选后剩余 {len(version_filtered)} 条成就数据")
            
            if not version_filtered:
                raise Exception(f"版本 {target_version} 没有找到任何成就数据")
                
            achievements = version_filtered
                
        except Exception as e:
            raise Exception(f"解析数据失败: {str(e)}")

        return achievements
    
    
    
    def fill_serial_numbers(self, achievements):
            """根据分类自动填充绝对编号和编号"""
            # 获取分类配置
            first_categories = self.first_categories
            second_categories = self.second_categories
            
            def get_sort_key(achievement):
                """获取排序键"""
                # 第一分类排序
                first_cat = achievement.get('第一分类', '')
                first_order = first_categories.get(first_cat, 999)

                # 第二分类排序
                second_cat = achievement.get('第二分类', '')
                first_cat_second = second_categories.get(first_cat, {})
                second_order = int(first_cat_second.get(second_cat, 999)) if second_cat in first_cat_second else 999

                # 版本号排序（浮点型正序）
                version_str = achievement.get('version', '0.0')
                try:
                    version = float(version_str)
                except ValueError:
                    version = 0.0

                # 原编号（用于保持相对稳定）
                original_id = achievement.get('serial_number', '99999999')

                return (first_order, second_order, version, original_id)
            
            # 按新规则排序
            sorted_achievements = sorted(achievements, key=get_sort_key)
            # 用于跟踪每个分类组合的当前序号
            current_numbers = {}
            
            # 为每个成就分配编号
            for achievement in sorted_achievements:
                first_cat = achievement.get('第一分类', '')
                second_cat = achievement.get('第二分类', '')
                
                if not first_cat or not second_cat:
                    achievement['serial_number'] = ''
                    continue
                
                # 获取第一分类
                first_category_detail = first_cat
                first_category = self.get_first_category(first_category_detail)
                
                # 获取第二分类后缀
                suffix = self.get_second_category_suffix(first_category, second_cat)
                
                # 获取第一分类排序号
                first_category_order = self.first_categories.get(first_category, 1)
                
                # 生成完整前缀：第一分类(1位) + 第二分类后缀(补齐到3位)
                # 确保第二分类后缀至少3位，不足前面补0
                suffix_padded = f"{int(suffix):03d}"
                full_prefix = f"{first_category_order}{suffix_padded}"
                
                # 获取当前序号
                category_key = (first_cat, second_cat)
                current_num = current_numbers.get(category_key, 1)
                
                # 生成编号：4位分类码 + 4位序号
                achievement['serial_number'] = f"{full_prefix}{current_num:04d}"
                
                # 更新序号
                current_numbers[category_key] = current_num + 1
            
            return sorted_achievements
    
    def get_first_category(self, first_category_detail):
        """获取第一分类，如果不存在则智能分配并添加到配置"""
        # 如果分类已存在，直接返回
        if first_category_detail in self.first_categories:
            return first_category_detail
        
        # 为新分类分配排序号（当前最大排序号+1）
        max_order = max(self.first_categories.values()) if self.first_categories else 0
        new_order = max_order + 1
        
        # 添加新分类到配置
        self.first_categories[first_category_detail] = new_order
        self.second_categories[first_category_detail] = {}
        
        # 保存配置
        self.save_category_config()
        
        return first_category_detail
    
    def get_second_category_suffix(self, first_category, second_category):
        """获取第二分类后缀，如果不存在则智能分配"""
        # 获取该第一分类下的第二分类配置
        category_config = self.second_categories.get(first_category, {})
        
        # 如果已存在，返回现有后缀
        if second_category in category_config:
            return category_config[second_category]
        
        # 智能分配新后缀
        existing_suffixes = set()
        for suffix in category_config.values():
            try:
                existing_suffixes.add(int(suffix))
            except (ValueError, TypeError):
                pass
        
        # 找到最小的未使用后缀
        new_suffix = 10
        while new_suffix in existing_suffixes:
            new_suffix += 10
        
        # 保存新配置
        category_config[second_category] = str(new_suffix)
        self.second_categories[first_category] = category_config
        
        # 保存到配置文件
        self.save_category_config()
        
        return str(new_suffix)
    
    def save_category_config(self):
        """保存分类配置到文件"""
        updated_config = {
            "first_categories": self.first_categories,
            "second_categories": self.second_categories
        }
        from core.config import config
        config.save_category_config(updated_config)
        print("[INFO] 分类配置已保存")


class CrawlerThread(QThread):
    """爬虫线程"""
    
    def __init__(self, crawler):
        super().__init__()
        self.crawler = crawler
    
    def run(self):
        self.crawler.crawl()


class AchievementTable(QTableWidget):
    """成就表格（爬虫专用，不带状态功能）"""
    
    def __init__(self):
        super().__init__()
        self.setup_table()
        
    def setup_table(self):
        """设置表格"""
        # 设置列
        headers = ['名称', '描述', '奖励', '版本', '隐藏', '第一分类', '第二分类']
        self.setColumnCount(len(headers))
        self.setHorizontalHeaderLabels(headers)
        
        # 设置表格属性
        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.horizontalHeader().setStretchLastSection(True)
        
        # 设置垂直表头（序号列）样式
        self.verticalHeader().setVisible(True)
        self.verticalHeader().setDefaultSectionSize(25)
        self.verticalHeader().setMinimumWidth(40)
        
        # 去掉选中框和焦点
        self.setShowGrid(False)
        self.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.setSelectionMode(QTableWidget.SelectionMode.NoSelection)  # 完全禁止选择
        # 动态获取主题颜色
        from core.styles import BaseStyles
        from core.config import config
        table_style = BaseStyles.get_text_input_style(config.theme)
        self.setStyleSheet(table_style)
        
        # 设置列宽
        self.setColumnWidth(0, 200)  # 名称
        self.setColumnWidth(1, 300)  # 描述
        self.setColumnWidth(2, 100)  # 奖励
        self.setColumnWidth(3, 100)  # 版本
        self.setColumnWidth(4, 80)   # 隐藏
        self.setColumnWidth(5, 120)  # 第一分类
        self.setColumnWidth(6, 120)  # 第二分类
        
        # 启用鼠标追踪以支持悬浮提示
        self.setMouseTracking(True)
        
    def load_data(self, achievements):
        """加载数据"""
        self.setRowCount(len(achievements))
        self.achievements = achievements  # 保存数据引用
        
        for row, achievement in enumerate(achievements):
            # 名称（完整显示，但保留悬浮提示）
            name = achievement.get('名称', '')
            name_item = QTableWidgetItem(name)
            name_item.setToolTip(name)  # 悬浮显示完整名称
            
            # 设置字体加粗
            font = name_item.font()
            font.setBold(True)
            name_item.setFont(font)
            
            if achievement.get('是否隐藏') == '隐藏':
                name_item.setForeground(QColor(255, 165, 0))  # 橙黄色文字
            self.setItem(row, 0, name_item)

            # 描述（完整显示，但保留悬浮提示）
            desc = achievement.get('描述', '')
            desc_item = QTableWidgetItem(desc)
            desc_item.setToolTip(desc)  # 悬浮显示完整描述
            self.setItem(row, 1, desc_item)

            # 奖励
            reward_item = QTableWidgetItem(achievement.get('奖励', ''))
            reward_text = achievement.get('奖励', '')
            if '20' in reward_text:
                reward_item.setForeground(QColor(255, 107, 53))  # 橙色
            elif '10' in reward_text:
                reward_item.setForeground(QColor(78, 205, 196))  # 青色
            elif '5' in reward_text:
                reward_item.setForeground(QColor(69, 183, 209))  # 蓝色
            self.setItem(row, 2, reward_item)

            # 版本
            self.setItem(row, 3, QTableWidgetItem(achievement.get('版本', '')))

            # 隐藏
            hidden_item = QTableWidgetItem("是" if achievement.get('是否隐藏') == '隐藏' else "否")
            if achievement.get('是否隐藏') == '隐藏':
                hidden_item.setForeground(QColor(255, 165, 0))  # 橙黄色文字
            self.setItem(row, 4, hidden_item)

            # 第一分类
            self.setItem(row, 5, QTableWidgetItem(achievement.get('第一分类', '')))

            # 第二分类
            self.setItem(row, 6, QTableWidgetItem(achievement.get('第二分类', '')))

        # 不调整列宽，保持设定的宽度
    



class CrawlTab(QWidget):
    """数据爬取标签页"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.achievements = []
        self.crawler_thread = None
        # 本地数据文件路径
        # 保存数据到JSON文件
        from core.config import get_resource_path
        data_dir = get_resource_path("resources")
        data_dir.mkdir(exist_ok=True)
        self.data_file = str(data_dir / "wuthering_waves_achievements.json")
        self.init_ui()
        
        # 配置信息
        # 从配置中读取认证信息
        self._load_auth_config()
        
        # 监听配置变化信号
        signal_bus.settings_changed.connect(self._on_settings_changed)
        
        # 监听主题切换信号
        signal_bus.theme_changed.connect(self._on_theme_changed)
        
        # 不再自动加载本地数据,避免显示旧数据
        # self.load_local_data()
        
    def init_ui(self):
        """初始化UI"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(15)
        
        # 控制面板
        control_group = QGroupBox("控制面板")
        control_layout = QHBoxLayout(control_group)
        
        self.version_input = QLineEdit()
        self.version_input.setPlaceholderText("输入版本号（如：1.0, 1.1, 2.0）")
        self.version_input.setMaximumWidth(150)
        # 添加输入验证器，只允许数字和小数点
        from PySide6.QtGui import QRegularExpressionValidator
        from PySide6.QtCore import QRegularExpression
        version_validator = QRegularExpressionValidator(QRegularExpression(r"^\d*\.?\d*$"))
        self.version_input.setValidator(version_validator)
        # 连接焦点失去信号，用于自动格式化
        self.version_input.editingFinished.connect(self.format_version_input)
        control_layout.addWidget(QLabel("版本:"))
        control_layout.addWidget(self.version_input)
        
        self.crawl_btn = QPushButton("开始爬取")
        self.crawl_btn.setStyleSheet(get_button_style(config.theme))
        self.crawl_btn.clicked.connect(self.start_crawling)
        self.crawl_btn.setMaximumWidth(100)
        control_layout.addWidget(self.crawl_btn)
        
        self.merge_btn = QPushButton("确认覆盖")
        self.merge_btn.setStyleSheet(get_button_style(config.theme))
        self.merge_btn.clicked.connect(self.merge_to_manage)
        self.merge_btn.setEnabled(False)
        self.merge_btn.setMaximumWidth(100)
        control_layout.addWidget(self.merge_btn)
        
        self.wiki_btn = QPushButton("打开WIKI")
        self.wiki_btn.setStyleSheet(get_button_style(config.theme))
        self.wiki_btn.clicked.connect(self.open_wiki)
        self.wiki_btn.setMaximumWidth(100)
        control_layout.addWidget(self.wiki_btn)
        
        self.clear_cache_btn = QPushButton("清除缓存")
        self.clear_cache_btn.setStyleSheet(get_button_style(config.theme))
        self.clear_cache_btn.clicked.connect(self.clear_cache)
        self.clear_cache_btn.setMaximumWidth(100)
        control_layout.addWidget(self.clear_cache_btn)
        
        self.export_template_btn = QPushButton("导出范本")
        self.export_template_btn.setStyleSheet(get_button_style(config.theme))
        self.export_template_btn.clicked.connect(self.export_excel_template)
        self.export_template_btn.setMaximumWidth(100)
        control_layout.addWidget(self.export_template_btn)
        
        self.import_excel_btn = QPushButton("导入Excel")
        self.import_excel_btn.setStyleSheet(get_button_style(config.theme))
        self.import_excel_btn.clicked.connect(self.import_excel)
        self.import_excel_btn.setMaximumWidth(100)
        control_layout.addWidget(self.import_excel_btn)
        
        self.export_btn = QPushButton("导出JSON")
        self.export_btn.setStyleSheet(get_button_style(config.theme))
        self.export_btn.clicked.connect(self.export_json)
        self.export_btn.setEnabled(False)
        self.export_btn.setMaximumWidth(100)
        control_layout.addWidget(self.export_btn)
        
        self.export_excel_btn = QPushButton("导出Excel")
        self.export_excel_btn.setStyleSheet(get_button_style(config.theme))
        self.export_excel_btn.clicked.connect(self.export_excel)
        self.export_excel_btn.setEnabled(False)
        self.export_excel_btn.setMaximumWidth(100)
        control_layout.addWidget(self.export_excel_btn)
        
        control_layout.addStretch()
        layout.addWidget(control_group)
        
        # 数据表格
        self.table = AchievementTable()
        layout.addWidget(self.table)
        
        # 初始日志
        print("[INFO] 数据爬取标签页已初始化")
    
    def _load_auth_config(self):
        """从配置中加载认证信息"""
        self.devcode, self.token = config.get_auth_data()
    
    def _on_settings_changed(self, settings_data):
        """配置变化时的处理"""
        if 'devcode' in settings_data or 'token' in settings_data:
            self._load_auth_config()
            print("[INFO] 认证配置已更新")
    
    def _on_theme_changed(self, theme):
        """主题切换时更新样式"""
        # 更新按钮样式
        if hasattr(self, 'crawl_btn'):
            self.crawl_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'merge_btn'):
            self.merge_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'wiki_btn'):
            self.wiki_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'clear_cache_btn'):
            self.clear_cache_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'export_template_btn'):
            self.export_template_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'import_excel_btn'):
            self.import_excel_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'export_btn'):
            self.export_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'export_excel_btn'):
            self.export_excel_btn.setStyleSheet(get_button_style(theme))
    
    def format_version_input(self):
        """自动格式化版本号输入"""
        text = self.version_input.text().strip()
        
        # 如果输入的是整数，自动添加.0
        if text and text.isdigit():
            self.version_input.setText(f"{text}.0")
    
    def start_crawling(self):
        """开始爬取"""
        target_version = self.version_input.text().strip()
        
        # 检查是否输入了版本号
        if not target_version:
            self.show_notification("请输入版本号")
            return
        
        self.crawl_btn.setEnabled(False)
        
        # 检查认证信息是否完整
        if not self.devcode or not self.token:
            self.crawl_btn.setEnabled(True)
            self.show_notification("请先在设置中配置认证信息")
            return
        
        # 创建爬虫对象和线程
        crawler = AchievementCrawler(target_version)
        crawler.progress.connect(self.update_progress)
        crawler.finished.connect(self.on_crawl_finished)
        crawler.error.connect(self.on_crawl_error)
        
        self.crawler_thread = CrawlerThread(crawler)
        self.crawler_thread.start()
        
        print(f"[INFO] 开始爬取成就数据，版本: {target_version or '全部'}")
    
    def update_progress(self, message):
        """更新进度"""
        print(f"[INFO] {message}")
    
    def on_crawl_finished(self, achievements):
        """爬取完成"""
        self.achievements = achievements
        self.table.load_data(achievements)
        self.export_btn.setEnabled(True)
        self.export_excel_btn.setEnabled(True)
        self.crawl_btn.setEnabled(True)
        
        self.show_notification(f"爬取完成，共获取 {len(achievements)} 条成就数据")
        
        # 更新配置中的默认输出文件名（包含版本）
        target_version = self.version_input.text().strip()
        if not target_version and achievements:
            # 如果输入框没有版本，尝试从数据中获取
            versions = set()
            for achievement in achievements:
                ver = achievement.get('版本', '')
                if ver:
                    versions.add(ver)
            
            if versions:
                # 如果有多个版本，显示版本范围
                if len(versions) == 1:
                    target_version = list(versions)[0]
                else:
                    # 排序版本并获取范围
                    sorted_versions = sorted(versions, key=lambda x: float(x) if x.replace('.', '').isdigit() else 0)
                    target_version = f"{sorted_versions[0]}-{sorted_versions[-1]}"
        
        if target_version:
            try:
                from core.config import config
                # 确保crawl_settings属性存在
                if not hasattr(config, 'crawl_settings'):
                    config.crawl_settings = {}
                
                config.crawl_settings["default_output_file"] = f"鸣潮v{target_version}爬取数据.json"
                
                # 确保save_config方法存在
                if hasattr(config, 'save_config'):
                    config.save_config()
                    print(f"[INFO] 已更新默认输出文件名为: 鸣潮v{target_version}爬取数据.json")
                else:
                    print("[WARNING] config对象缺少save_config方法")
            except Exception as e:
                print(f"[ERROR] 更新配置失败: {str(e)}")
        
        print(f"[SUCCESS] 爬取完成，共 {len(achievements)} 条数据")
        
        # 启用确认覆盖按钮
        self.merge_btn.setEnabled(True)
    
    def on_crawl_error(self, error_message):
        """爬取出错"""
        self.crawl_btn.setEnabled(True)
        self.show_notification(f"爬取失败: {error_message}")
        print(f"[ERROR] 爬取失败: {error_message}")
    
    def merge_to_manage(self):
        """确认覆盖到成就管理"""
        if not self.achievements:
            print("[WARNING] 没有数据可以覆盖")
            return
        
        # 添加确认对话框
        from core.custom_message_box import CustomMessageBox
        reply = CustomMessageBox.question(
            self, 
            "确认添加", 
            f"确定要将新爬取的 {len(self.achievements)} 条成就数据添加到现有数据中吗？\n仅添加不存在的成就，已存在的成就将保持不变！",
            ("确定", "取消")
        )
        
        if reply != CustomMessageBox.Yes:
            print("[INFO] 用户取消了覆盖操作")
            return
        
        # 获取当前管理标签页的数据
        # 尝试多种方式获取main_window
        main_window = None
        
        # 方法1: 通过parent链
        try:
            parent = self.parent()
            if parent:
                tab_widget = parent.parent()
                if tab_widget:
                    main_window = tab_widget.parent()
        except:
            pass
        
        # 方法2: 通过全局查找
        if not main_window or not hasattr(main_window, 'manage_tab'):
            try:
                from core.main_window import TemplateMainWindow
                for widget in QApplication.topLevelWidgets():
                    if isinstance(widget, TemplateMainWindow):
                        main_window = widget
                        break
                    # 即使不是TemplateMainWindow类型，如果有manage_tab属性也尝试使用
                    elif hasattr(widget, 'manage_tab'):
                        main_window = widget
                        break
            except:
                pass
        
        if not main_window or not hasattr(main_window, 'manage_tab'):
            print("[ERROR] 找不到管理标签页")
            return
        
        manage_tab = main_window.manage_tab
        if not manage_tab:
            print("[ERROR] 管理标签页为空")
            return
        
        current_achievements = manage_tab.manager.achievements
        
        # 创建名称+描述的组合键，用于精确判断重复
        # 去掉描述末尾符号后再比较
        import re
        def clean_description(desc):
            """去掉描述末尾的标点符号"""
            if not desc:
                return desc
            # 去掉末尾的标点符号：。，；：！？、
            return re.sub(r'[.,…。，；：！？、]+$', '', desc).strip()
        
        current_achievements_keys = set()
        for a in current_achievements:
            name = a.get('名称', '')
            desc = clean_description(a.get('描述', ''))
            key = (name, desc)  # 使用元组作为组合键
            current_achievements_keys.add(key)
        
        # 仅筛选出不存在的成就进行添加
        to_add = []     # 需要新增的成就
        
        for achievement in self.achievements:
            name = achievement.get('名称', '')
            desc = clean_description(achievement.get('描述', ''))
            key = (name, desc)
            
            if key not in current_achievements_keys:
                # 新成就（名称+描述组合不存在）
                to_add.append(achievement)
        
        if not to_add:
            print("[INFO] 所有成就已存在，无需添加")
            self.show_notification("所有成就已存在，无需添加")
            return
        
        # 获取分类配置
        from core.config import config
        category_config = config.load_category_config()
        first_categories = category_config.get("first_categories", {})
        second_categories = category_config.get("second_categories", {})
        
        # 复制一份配置用于更新
        updated_first_categories = first_categories.copy()
        updated_second_categories = {}
        for key, value in second_categories.items():
            updated_second_categories[key] = value.copy()
        
        # 标记是否有新的分类需要保存
        has_new_categories = False
        
        # 检查新成就中是否有新的分类
        for achievement in to_add:
            first_cat = achievement.get('第一分类', '')
            second_cat = achievement.get('第二分类', '')
            
            if first_cat and second_cat:
                # 智能处理第一分类（如果不存在则分配新排序）
                if first_cat not in updated_first_categories:
                    max_order = max(updated_first_categories.values()) if updated_first_categories else 0
                    updated_first_categories[first_cat] = max_order + 1
                    updated_second_categories[first_cat] = {}
                    has_new_categories = True
                    print(f"[INFO] 发现新第一分类 '{first_cat}'，分配排序: {max_order + 1}")
                
                # 智能处理第二分类（如果不存在则分配新后缀）
                if first_cat not in updated_second_categories:
                    updated_second_categories[first_cat] = {}
                
                if second_cat not in updated_second_categories[first_cat]:
                    # 找到该第一分类下最小的未使用后缀
                    existing_suffixes = set()
                    for suffix in updated_second_categories[first_cat].values():
                        try:
                            existing_suffixes.add(int(suffix))
                        except (ValueError, TypeError):
                            pass
                    
                    new_suffix = 10
                    while new_suffix in existing_suffixes:
                        new_suffix += 10
                    
                    updated_second_categories[first_cat][second_cat] = str(new_suffix)
                    has_new_categories = True
                    print(f"[INFO] 发现新第二分类 '{first_cat} - {second_cat}'，分配后缀: {new_suffix}")
        
        # 先保存新的分类配置（如果有新分类）
        if has_new_categories:
            updated_config = {
                "first_categories": updated_first_categories,
                "second_categories": updated_second_categories
            }
            config.save_category_config(updated_config)
            print("[INFO] 已更新分类配置，新增的分类已自动分配排序和后缀")        # 合并数据：现有成就 + 新增的成就
        all_achievements = current_achievements + to_add
        
        # 使用ManageTab的智能重新编码方法来重新生成编号和绝对编号
        all_achievements = manage_tab._smart_reencode_achievements(all_achievements)
        
        # 直接更新管理器的数据，而不是调用load_data
        manage_tab.manager.achievements = all_achievements
        manage_tab.manager.filtered_achievements = all_achievements.copy()
        
        # 重新编码所有用户的存档数据，确保编号同步
        print("[INFO] 正在重新编码用户存档数据...")
        if config.reencode_all_user_progress():
            print("[SUCCESS] 用户存档数据已同步更新")
        else:
            print("[ERROR] 用户存档数据更新失败")
        
        # 发送分类配置更新信号（如果有新分类）
        if has_new_categories:
            from core.signal_bus import signal_bus
            signal_bus.category_config_updated.emit()
        
        print(f"[SUCCESS] 已新增 {len(to_add)} 条成就，总计 {len(all_achievements)} 条成就数据")
        
        # 显示多个通知
        if has_new_categories:
            # 先显示新分类通知
            self.show_notification("发现新分类，已自动分配排序。建议到设置→分类管理中手动调整顺序。")
            # 延迟0.5秒后显示成功通知
            from PySide6.QtCore import QTimer
            QTimer.singleShot(500, lambda: self.show_notification(f"成功新增 {len(to_add)} 条成就，总计 {len(all_achievements)} 条成就数据"))
        else:
            # 没有新分类时直接显示成功通知
            self.show_notification(f"成功新增 {len(to_add)} 条成就，总计 {len(all_achievements)} 条成就数据")
        
        # 更新爬虫页面的表格显示
        manage_tab.manager_table.load_data(all_achievements)
        
        # 更新主窗口成就管理页面的数据
        manage_tab.manager.achievements = all_achievements
        manage_tab.manager.filtered_achievements = all_achievements.copy()
        
        # 更新主窗口成就管理页面的表格
        if hasattr(manage_tab, 'achievement_table'):
            manage_tab.achievement_table.load_data(all_achievements)
        
        # 更新筛选器
        manage_tab.update_filters()
        
        # 更新统计
        manage_tab.update_statistics()
        
        # 保存到JSON
        manage_tab.save_to_json()
        
        # 禁用确认覆盖按钮
        self.merge_btn.setEnabled(False)
        
        # 切换到管理标签页
        if hasattr(main_window, 'tab_widget'):
            main_window.tab_widget.setCurrentIndex(0)  # 成就管理现在是第一个标签页
    
    def show_notification(self, message):
        """显示右上角自动关闭的提示"""
        from PySide6.QtWidgets import QLabel, QVBoxLayout
        from PySide6.QtCore import QTimer, Qt
        
        # 创建提示窗口
        notification = QWidget()
        notification.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.Tool | Qt.WindowType.WindowStaysOnTopHint)
        notification.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        notification.setAttribute(Qt.WidgetAttribute.WA_DeleteOnClose)
        
        layout = QVBoxLayout(notification)
        layout.setContentsMargins(0, 0, 0, 0)
        
        label = QLabel(message)
        
        # 使用统一的通知样式
        from core.styles import get_notification_style
        label.setStyleSheet(get_notification_style(config.theme))
        label.setWordWrap(True)
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(label)
        
        # 获取主窗口位置
        main_window = None
        try:
            from core.main_window import TemplateMainWindow
            for widget in QApplication.topLevelWidgets():
                if isinstance(widget, TemplateMainWindow):
                    main_window = widget
                    break
        except:
            pass
        
        # 设置提示大小和位置
        notification_width = 300
        notification_height = 60
        notification.setFixedSize(notification_width, notification_height)
        
        if main_window:
            # 相对于主窗口右上角
            main_pos = main_window.pos()
            main_size = main_window.size()
            x = main_pos.x() + main_size.width() - notification_width - 20
            y = main_pos.y() + 60
        else:
            # 屏幕右上角
            from PySide6.QtGui import QGuiApplication
            screen = QGuiApplication.primaryScreen().geometry()
            x = screen.width() - notification_width - 20
            y = 60
        
        # 计算垂直位置，考虑现有通知
        if not hasattr(self, 'active_notifications'):
            self.active_notifications = []
        
        # 垂直偏移量：每个通知间隔10px
        vertical_offset = len(self.active_notifications) * 60
        notification.move(x, y + vertical_offset)
        notification.show()
        
        # 创建淡出定时器
        fade_timer = QTimer()
        fade_timer.setSingleShot(True)
        fade_timer.timeout.connect(lambda: self.fade_out_notification(notification))
        fade_timer.start(6000)  # 6秒后开始淡出
        
        # 存储引用以避免被垃圾回收
        self.active_notifications.append((notification, fade_timer))
    
    def fade_out_notification(self, notification):
        """淡出提示"""
        from PySide6.QtCore import QPropertyAnimation, QEasingCurve
        
        # 检查对象是否还存在
        try:
            if notification is None or not notification.isVisible():
                self._remove_notification_from_list(notification)
                self._reposition_notifications()
                return
        except RuntimeError:
            # 对象已被删除
            self._remove_notification_from_list(notification)
            self._reposition_notifications()
            return
        
        # 创建透明度动画
        fade_animation = QPropertyAnimation(notification, b"windowOpacity")
        fade_animation.setDuration(500)
        fade_animation.setStartValue(1.0)
        fade_animation.setEndValue(0.0)
        fade_animation.setEasingCurve(QEasingCurve.Type.OutQuad)
        
        # 动画完成后删除提示
        fade_animation.finished.connect(lambda: self._cleanup_notification(notification))
        fade_animation.start()
        
        # 保存动画引用防止被垃圾回收
        if not hasattr(self, 'active_animations'):
            self.active_animations = []
        self.active_animations.append(fade_animation)
        
        # 从活动通知列表中移除
        self._remove_notification_from_list(notification)
        # 重新定位剩余通知
        self._reposition_notifications()
    
    def _remove_notification_from_list(self, notification):
        """从活动通知列表中移除通知"""
        if hasattr(self, 'active_notifications'):
            self.active_notifications = [(n, t) for n, t in self.active_notifications if n != notification]
    
    def _reposition_notifications(self):
        """重新定位所有活动通知"""
        if not hasattr(self, 'active_notifications'):
            return
            
        # 获取主窗口位置
        main_window = None
        try:
            from core.main_window import TemplateMainWindow
            for widget in QApplication.topLevelWidgets():
                if isinstance(widget, TemplateMainWindow):
                    main_window = widget
                    break
                elif hasattr(widget, 'manage_tab'):
                    main_window = widget
                    break
        except:
            pass
        
        # 计算基础位置
        if main_window:
            # 相对于主窗口右上角
            main_pos = main_window.pos()
            main_size = main_window.size()
            x = main_pos.x() + main_size.width() - 320 - 20  # 320是通知宽度
            y = main_pos.y() + 60
        else:
            # 屏幕右上角
            from PySide6.QtGui import QGuiApplication
            screen = QGuiApplication.primaryScreen().geometry()
            x = screen.width() - 320 - 20
            y = 60
        
        # 重新定位每个通知
        for i, (notification, timer) in enumerate(self.active_notifications):
            if notification and notification.isVisible():
                vertical_offset = i * 60
                notification.move(x, y + vertical_offset)
    
    def open_wiki(self):
        """打开库街区Wiki成就页面"""
        import webbrowser
        wiki_url = "https://wiki.kurobbs.com/mc/item/1220879855033786368?wkFrom=home&wkFromLabel=%E9%A6%96%E9%A1%B5%E5%BF%AB%E6%8D%B7%E5%AF%BC%E8%88%AA"
        webbrowser.open(wiki_url)
        print(f"[INFO] 已打开Wiki页面: {wiki_url}")
    
    def clear_cache(self):
        """清除本地缓存的网页文件"""
        from core.custom_message_box import CustomMessageBox
        from core.config import get_resource_path
        
        cache_file = get_resource_path("resources") / "achievement_cache.json"
        
        if cache_file.exists():
            reply = CustomMessageBox.question(
                self, 
                "确认清除", 
                "确定要清除本地缓存的网页文件吗？",
                ("确定", "取消")
            )
            
            if reply == CustomMessageBox.Yes:
                try:
                    cache_file.unlink()
                    print("[INFO] 缓存文件已删除")
                    self.show_notification("缓存已清除")
                except Exception as e:
                    print(f"[ERROR] 清除缓存失败: {str(e)}")
                    self.show_notification(f"清除缓存失败: {str(e)}")
        else:
            print("[INFO] 没有找到缓存文件")
            self.show_notification("没有缓存文件")
    
    def _cleanup_notification(self, notification):
        """清理通知"""
        try:
            if notification and hasattr(notification, 'close'):
                notification.close()
            if notification and hasattr(notification, 'deleteLater'):
                notification.deleteLater()
            # 清理动画引用
            if hasattr(self, 'active_animations'):
                self.active_animations = [anim for anim in self.active_animations if anim and anim.state() != anim.State.Stopped]
        except:
            pass
    
        def on_crawl_error(self, error_msg):
    
            """爬取错误"""
    
            self.crawl_btn.setEnabled(True)
    
            print(f"[ERROR] 爬取失败: {error_msg}")
    
            self.show_notification(f"爬取失败: {error_msg}")
    
    
    
    def export_excel_template(self):
        """导出Excel范本"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出Excel范本", "成就数据导入范本.xlsx", "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                self.create_excel_template(file_path)
            except Exception as e:
                print(f"[ERROR] 导出范本失败: {str(e)}")
                self.show_notification(f"导出范本失败: {str(e)}")
    
    def create_excel_template(self, file_path):
        """创建Excel范本文件"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 创建工作簿
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "成就数据导入范本"
            
            # 定义列信息
            columns_info = [
                {'name': '名称', 'required': '是', 'description': '成就名称，必须列。系统只自动去除"「隐藏成就」"标志，其他隐藏字眼需手动清理', 'example': '「隐藏成就」我们相信漂泊者'},
                {'name': '描述', 'required': '否', 'description': '成就描述，非必须列', 'example': '重建拉海洛的全部路网'},
                {'name': '版本', 'required': '是', 'description': '版本号，必须列，整数自动补.0', 'example': '3 或 3.0'},
                {'name': '奖励', 'required': '是', 'description': '奖励内容，必须列，纯数字自动加"星声*"', 'example': '10 或 星声*10'},
                {'name': '是否隐藏', 'required': '否', 'description': '是否隐藏，非必须列，建议手动填写', 'example': '隐藏 或 留空'},
                {'name': '第一分类', 'required': '否', 'description': '第一分类，非必须列，不提供则根据第二分类自动获取', 'example': '索拉漫行'},
                {'name': '第二分类', 'required': '是', 'description': '第二分类，必须列', 'example': '索拉的大地·拉海洛'}
            ]
            
            # 设置表头
            headers = ['列名', '是否必须', '说明', '示例数据']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 填充列说明信息
            for row, col_info in enumerate(columns_info, 2):
                # 列名
                cell = sheet.cell(row=row, column=1, value=col_info['name'])
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
                
                # 是否必须
                cell = sheet.cell(row=row, column=2, value=col_info['required'])
                cell.font = Font(bold=True, color="FF0000" if col_info['required'] == '是' else "000000")
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
                
                # 说明
                cell = sheet.cell(row=row, column=3, value=col_info['description'])
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
                
                # 示例数据
                cell = sheet.cell(row=row, column=4, value=col_info['example'])
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 调整列宽
            sheet.column_dimensions[get_column_letter(1)].width = 15  # 列名
            sheet.column_dimensions[get_column_letter(2)].width = 10  # 是否必须
            sheet.column_dimensions[get_column_letter(3)].width = 50  # 说明
            sheet.column_dimensions[get_column_letter(4)].width = 25  # 示例数据
            
            # 添加数据示例区域标题（考虑新增的分隔线和提示）
            example_title_row = len(columns_info) + 5
            title_cell = sheet.cell(row=example_title_row, column=1, value="📋 数据示例区域（导入时删除当前行及以上所有内容行，只使用此区域数据）：")
            title_cell.font = Font(bold=True, size=12, color="FF0000")
            title_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            # 合并标题单元格
            sheet.merge_cells(start_row=example_title_row, start_column=1, 
                            end_row=example_title_row, end_column=7)
            
            # 添加示例数据表头
            example_headers_row = example_title_row + 1
            example_headers = ['名称', '描述', '奖励', '版本', '是否隐藏', '第一分类', '第二分类']
            for col, header in enumerate(example_headers, 1):
                cell = sheet.cell(row=example_headers_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                   top=Side(style='thin'), bottom=Side(style='thin'))
            
            # 添加示例数据行
            example_data = [
                ['我们相信漂泊者', '重建拉海洛的全部路网。', '10', '3', '', '索拉漫行', '索拉的大地·拉海洛'],
                ['请勿剐蹭', '在拉海洛路网上与车辆发生碰撞。', '5', '3', '隐藏', '索拉漫行', '索拉的大地·拉海洛'],
                ['心无妄虑', '完成「全息战略·同步」的「无妄者I」。', '星声*5', '3.0', '', '铿锵刃鸣', '来自深塔·二']
            ]
            
            for row_idx, row_data in enumerate(example_data, example_headers_row + 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                    # 为示例数据区域添加浅绿色背景，使其更醒目
                    cell.fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                       top=Side(style='thin'), bottom=Side(style='thin'))
                    if col_idx == 1:  # 名称列加粗
                        cell.font = Font(bold=True)
            
            # 调整示例数据区域列宽
            for col in range(1, 8):
                sheet.column_dimensions[get_column_letter(col)].width = 20
            
            
            
            # 保存文件
            workbook.save(file_path)
            print(f"[SUCCESS] Excel范本已导出到: {file_path}")
            self.show_notification("范本导出成功")
            
        except Exception as e:
            print(f"[ERROR] 创建Excel范本失败: {str(e)}")
            raise Exception(f"创建Excel范本失败: {str(e)}")

    def import_excel(self):
        """导入Excel文件"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "导入Excel文件", "", "Excel Files (*.xlsx *.xls)"
        )
        
        if file_path:
            try:
                self.import_from_excel(file_path)
            except Exception as e:
                print(f"[ERROR] 导入失败: {str(e)}")
                self.show_notification(f"导入失败: {str(e)}")
    
    def import_from_excel(self, excel_path):
        """从Excel文件导入数据并进行清洗"""
        try:
            from openpyxl import load_workbook
            
            # 读取Excel文件
            print(f"[INFO] 正在读取Excel文件: {excel_path}")
            workbook = load_workbook(excel_path)
            sheet = workbook.active
            
            # 获取表头
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            # 检查必要的列
            required_columns = ['名称', '第二分类']
            missing_columns = [col for col in required_columns if col not in headers]
            if missing_columns:
                raise Exception(f"缺少必要的列: {', '.join(missing_columns)}")
            
            # 创建列名到索引的映射
            col_index = {header: idx for idx, header in enumerate(headers)}
            
            # 数据清洗和转换
            print(f"[INFO] 开始数据清洗...")
            cleaned_achievements = []
            
            # 加载分类配置
            try:
                # 确保config对象已正确初始化
                if not hasattr(config, 'load_category_config'):
                    print("[ERROR] config对象缺少load_category_config方法")
                    raise Exception("配置对象未正确初始化")
                
                category_config = config.load_category_config()
                if not isinstance(category_config, dict):
                    print(f"[ERROR] category_config不是字典类型: {type(category_config)}")
                    category_config = {}
                
                first_categories = category_config.get("first_categories", {})
                second_categories = category_config.get("second_categories", {})
                
                # 确保返回的是字典
                if not isinstance(first_categories, dict):
                    first_categories = {}
                if not isinstance(second_categories, dict):
                    second_categories = {}
                    
            except Exception as e:
                print(f"[ERROR] 加载分类配置失败: {str(e)}")
                # 使用默认配置
                first_categories = {}
                second_categories = {}
                print("[INFO] 使用默认分类配置")
            
            # 创建第二分类到第一分类的映射
            first_category_map = {}
            for first_cat, second_cats in second_categories.items():
                for second_cat in second_cats:
                    first_category_map[second_cat] = first_cat
            
            # 从第二行开始读取数据
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                if not any(cell.value for cell in row):
                    continue  # 跳过空行
                
                achievement = {}
                
                # 1. 名称列：去除「隐藏成就」
                name_value = row[col_index['名称']].value
                name = str(name_value).strip() if name_value else ''
                if '「隐藏成就」' in name:
                    name = name.replace('「隐藏成就」', '').strip()
                achievement['名称'] = name
                
                # 2. 描述列
                desc_value = row[col_index['描述']].value if '描述' in col_index else ''
                description = str(desc_value).strip() if desc_value else ''
                achievement['描述'] = description
                
                # 3. 版本列：智能处理小数
                version_value = row[col_index['版本']].value if '版本' in col_index else ''
                version = str(version_value).strip() if version_value else ''
                if version:
                    # 检查是否已经包含小数点
                    if '.' in version:
                        # 已经有小数点，保持原样
                        pass
                    else:
                        # 没有小数点，补充.0
                        version = f"{version}.0"
                achievement['版本'] = version
                
                # 4. 奖励列：纯数字拼接"星声*"
                reward_value = row[col_index['奖励']].value if '奖励' in col_index else ''
                reward = str(reward_value).strip() if reward_value else ''
                if reward.isdigit():
                    reward = f"星声*{reward}"
                achievement['奖励'] = reward
                
                # 5. 是否隐藏列：简化判断，只判断是否包含"隐藏"
                if '是否隐藏' in col_index:
                    hidden_value = row[col_index['是否隐藏']].value
                    is_hidden = str(hidden_value).strip() if hidden_value else ''
                    achievement['是否隐藏'] = '隐藏' if '隐藏' in is_hidden else ''
                else:
                    # 根据名称判断
                    achievement['是否隐藏'] = '隐藏' if '隐藏' in name else ''
                
                # 6. 第二分类列：必须有
                second_category_value = row[col_index['第二分类']].value
                second_category = str(second_category_value).strip() if second_category_value else ''
                if not second_category:
                    raise Exception(f"第{row_idx}行：第二分类不能为空")
                achievement['第二分类'] = second_category
                
                # 7. 第一分类列：如果没有提供，根据第二分类获取
                if '第一分类' in col_index:
                    first_category_value = row[col_index['第一分类']].value
                    first_category = str(first_category_value).strip() if first_category_value else ''
                    if first_category:
                        achievement['第一分类'] = first_category
                    else:
                        # 根据第二分类映射获取第一分类
                        first_category = first_category_map.get(second_category, '')
                        if first_category:
                            achievement['第一分类'] = first_category
                        else:
                            # 如果找不到对应的第一分类，收集所有缺失的分类
                            if not hasattr(self, 'missing_categories'):
                                self.missing_categories = set()
                            self.missing_categories.add(second_category)
                            continue  # 跳过这一行，继续检查下一行
                else:
                    # 根据第二分类映射获取第一分类
                    first_category = first_category_map.get(second_category, '')
                    if first_category:
                        achievement['第一分类'] = first_category
                    else:
                        # 如果找不到对应的第一分类，收集所有缺失的分类
                        if not hasattr(self, 'missing_categories'):
                            self.missing_categories = set()
                        self.missing_categories.add(second_category)
                        continue  # 跳过这一行，继续检查下一行
                
                cleaned_achievements.append(achievement)
            
            # 检查是否有缺失的分类
            if hasattr(self, 'missing_categories') and self.missing_categories:
                missing_list = sorted(list(self.missing_categories))
                missing_str = "、".join(missing_list)
                
                workbook.close()
                
                # 显示错误提示，中断导入
                error_msg = f"发现未配置的第二分类: {missing_str}\n\n"
                error_msg += "请在 设置→分类管理 中将这些分类添加到对应的第一分类下，然后重新导入。"
                
                print(f"[ERROR] 导入中断：发现未配置的分类: {missing_str}")
                print(f"[INFO] 请在 设置→分类管理 中添加这些分类后重新导入")
                
                # 显示错误提示
                from core.custom_message_box import CustomMessageBox
                CustomMessageBox.warning(self, "导入中断", error_msg)
                
                # 不更新数据，保持原状
                return
            
            workbook.close()
            
            # 更新数据
            self.achievements = cleaned_achievements
            self.table.load_data(cleaned_achievements)
            self.export_btn.setEnabled(True)
            self.export_excel_btn.setEnabled(True)
            self.merge_btn.setEnabled(True)
            
            print(f"[SUCCESS] 导入完成，共 {len(cleaned_achievements)} 条成就数据")
            
            # 检查是否有缺失的分类需要提示用户
            if hasattr(self, 'missing_categories') and self.missing_categories:
                missing_list = sorted(list(self.missing_categories))
                missing_str = "、".join(missing_list)
                print(f"[INFO] 发现未配置的第二分类: {missing_str}")
                print(f"[INFO] 请在 设置→分类管理 中将这些分类添加到对应的第一分类下")
                self.show_notification(f"导入成功！发现未配置分类: {missing_str}，请在设置→分类管理中添加")
            else:
                self.show_notification(f"导入成功，共 {len(cleaned_achievements)} 条成就数据")
            
        except Exception as e:
            print(f"[ERROR] 导入Excel失败: {str(e)}")
            raise Exception(f"导入Excel失败: {str(e)}")

    def export_json(self):
        """导出数据"""
        if not self.achievements:
            print("[WARNING] 没有数据可导出")
            return
        
        # 获取配置中的默认文件名
        from core.config import config
        default_filename = config.crawl_settings.get("default_output_file", "鸣潮成就数据.json")
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出JSON文件", default_filename, "JSON Files (*.json)"
        )
        
        if file_path:
            try:
                self.export_to_json(file_path)
            except Exception as e:
                print(f"[ERROR] 导出失败: {str(e)}")
    
    def export_to_json(self, json_path):
        """导出为全字段 JSON 格式"""
        try:
            export_data = []
            for achievement in self.achievements:
                export_data.append({
                    '绝对编号': achievement.get('绝对编号', ''),
                    '版本': achievement.get('版本', ''),
                    '第一分类': achievement.get('第一分类', ''),
                    '第二分类': achievement.get('第二分类', ''),
                    '编号': achievement.get('编号', ''),
                    '名称': achievement.get('名称', ''),
                    '描述': achievement.get('描述', ''),
                    '奖励': achievement.get('奖励', ''),
                    '是否隐藏': achievement.get('是否隐藏', ''),
                    '获取状态': achievement.get('获取状态', '')
                })
            
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            
            print(f"[SUCCESS] 全字段数据已导出到: {json_path}")
            print(f"[INFO] 包含 {len(export_data)} 条成就数据")
        except Exception as e:
            print(f"[ERROR] 导出 JSON 失败: {str(e)}")
    
    def export_excel(self):
        """导出Excel文件"""
        if not self.achievements:
            print("[WARNING] 没有数据可导出")
            show_notification(self, "没有数据可导出")
            return
        
        # 动态获取版本信息用于文件名
        version = self.version_input.text().strip()
        if not version and self.achievements:
            # 如果输入框没有版本，尝试从数据中获取
            versions = set()
            for achievement in self.achievements:
                ver = achievement.get('版本', '')
                if ver:
                    versions.add(ver)
            
            if versions:
                # 如果有多个版本，显示版本范围
                if len(versions) == 1:
                    version = list(versions)[0]
                else:
                    # 排序版本并获取范围
                    sorted_versions = sorted(versions, key=lambda x: float(x) if x.replace('.', '').isdigit() else 0)
                    version = f"{sorted_versions[0]}-{sorted_versions[-1]}"
        
        if version:
            default_filename = f"鸣潮v{version}爬取数据.xlsx"
        else:
            default_filename = "鸣潮爬取数据.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出Excel文件", default_filename, "Excel Files (*.xlsx)"
        )
        
        if file_path:
            try:
                self.export_to_excel(file_path)
            except Exception as e:
                print(f"[ERROR] 导出Excel失败: {str(e)}")
                self.show_notification(f"导出Excel失败: {str(e)}")
    
    def export_to_excel(self, excel_path):
        """导出为Excel格式"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            sheet = wb.active
            sheet.title = "成就数据"
            
            # 定义列顺序（与GUI表格保持一致）
            column_order = [
                '名称', '描述', '奖励', '版本', '是否隐藏', '第一分类', '第二分类'
            ]
            
            # 写入表头
            for col_idx, field_name in enumerate(column_order, 1):
                cell = sheet.cell(row=1, column=col_idx, value=field_name)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # 写入数据
            for row_idx, achievement in enumerate(self.achievements, start=2):
                for col_idx, field_name in enumerate(column_order, 1):
                    value = achievement.get(field_name, '')
                    
                    # 特殊处理名称列
                    if field_name == '名称' and value:
                        cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                        cell.font = Font(bold=True)
                        
                        # 隐藏成就用橙色
                        if achievement.get('是否隐藏') == '隐藏':
                            cell.font = Font(bold=True, color="FFA500")
                    else:
                        cell = sheet.cell(row=row_idx, column=col_idx, value=str(value))
                    
                    # 设置边框
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
            
            # 调整列宽
            column_widths = {
                '名称': 25, '描述': 40, '版本': 10, '奖励': 15, '是否隐藏': 10, 
                '第一分类': 15, '第二分类': 20
            }
            
            for col_idx, field_name in enumerate(column_order, 1):
                col_letter = get_column_letter(col_idx)
                sheet.column_dimensions[col_letter].width = column_widths.get(field_name, 15)
            
            # 保存文件
            wb.save(excel_path)
            
            print(f"[SUCCESS] Excel数据已导出到: {excel_path}")
            print(f"[INFO] 包含 {len(self.achievements)} 条成就数据")
            show_notification(self, f"成功导出 {len(self.achievements)} 条成就数据到Excel")
            
        except Exception as e:
            print(f"[ERROR] 导出Excel失败: {str(e)}")
            raise Exception(f"导出Excel失败: {str(e)}")
    
    def load_local_data(self):
        """加载本地保存的数据"""
        try:
            print(f"[DEBUG] 检查文件: {self.data_file}")
            print(f"[DEBUG] 文件是否存在: {os.path.exists(self.data_file)}")
            
            if os.path.exists(self.data_file):
                print(f"[INFO] 正在从 {self.data_file} 加载数据...")
                
                # 检查是否有 JSON 备份文件
                json_file = self.data_file.replace('.xlsx', '.json')
                if os.path.exists(json_file):
                    print(f"[INFO] 找到 JSON 备份文件，优先加载: {json_file}")
                    self.load_from_json(json_file)
                    return
                
                # 现在只支持JSON格式，Excel文件已不再使用
                print(f"[INFO] 本地数据文件不存在或格式不支持: {self.data_file}")
                print(f"[INFO] 请使用JSON格式的数据文件")
                
                print(f"[DEBUG] 转换后的数据条数: {len(self.achievements)}")
                
                self.table.load_data(self.achievements)
                self.export_btn.setEnabled(True)
                print(f"[INFO] 已加载本地数据: {len(self.achievements)} 条")
            else:
                print(f"[INFO] 本地数据文件不存在: {self.data_file}")
        except Exception as e:
            print(f"[WARNING] 加载本地数据失败: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def load_from_json(self, json_file):
        """从 JSON 文件加载数据"""
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                self.achievements = []
                
                # 直接使用中文键名,不需要映射
                for item in data:
                    achievement = {}
                    for key, value in item.items():
                        achievement[key] = value
                    
                    self.achievements.append(achievement)
                
                print(f"[INFO] 从 JSON 加载了 {len(self.achievements)} 条数据")
        except Exception as e:
            print(f"[ERROR] 加载 JSON 文件失败: {str(e)}")
    

    
    
    
    
    
    def save_local_data(self):
            """保存数据到本地文件（JSON）"""
            if not self.achievements:
                print("[WARNING] 没有数据可保存")
                return

            try:
                print(f"[DEBUG] 准备保存数据到: {self.data_file}")
                print(f"[DEBUG] 文件路径类型: {type(self.data_file)}")
                print(f"[DEBUG] 文件是否存在: {os.path.exists(self.data_file)}")

                # 确保目录存在
                data_dir = os.path.dirname(self.data_file)
                if not os.path.exists(data_dir):
                    print(f"[INFO] 创建目录: {data_dir}")
                    os.makedirs(data_dir, exist_ok=True)

                # 保存为 JSON 格式
                self.save_to_json(self.data_file)
                print("[INFO] 数据已保存为 JSON 格式")

            except Exception as e:
                print(f"[ERROR] 保存数据失败: {str(e)}")
                import traceback
                traceback.print_exc()
    
    def save_to_json(self, json_file):
        """保存为 JSON 格式"""
        try:
            # 准备数据，使用中文键名
            json_data = []
            for achievement in self.achievements:
                # 创建副本，确保数据可序列化
                item = {}
                
                # 按照指定顺序处理字段
                field_order = [
                    '版本',
                    '第一分类',
                    '第二分类',
                    '编号',
                    '名称',
                    '描述',
                    '奖励',
                    '是否隐藏'
                ]
                
                for chinese_key in field_order:
                    # 获取原始值
                    value = achievement.get(chinese_key, '')
                    
                    # 确保所有值都是字符串
                    if value is None:
                        item[chinese_key] = ''
                    else:
                        item[chinese_key] = str(value)
                
                json_data.append(item)
            
            # 保存到 JSON
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, ensure_ascii=False, indent=2)
            
            file_size = os.path.getsize(json_file)
            print(f"[SUCCESS] JSON 数据已保存到: {json_file}")
            print(f"[SUCCESS] JSON 文件大小: {file_size} 字节")
            print(f"[INFO] 保存了 {len(json_data)} 条成就数据")
            
        except Exception as e:
            print(f"[ERROR] 保存 JSON 失败: {str(e)}")
    
    
    
    def apply_theme(self, theme):
        """应用主题"""
        # 更新按钮样式
        if hasattr(self, 'crawl_btn'):
            self.crawl_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'merge_btn'):
            self.merge_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'export_template_btn'):
            self.export_template_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'import_excel_btn'):
            self.import_excel_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'export_btn'):
            self.export_btn.setStyleSheet(get_button_style(theme))
        if hasattr(self, 'export_excel_btn'):
            self.export_excel_btn.setStyleSheet(get_button_style(theme))
        
        # 更新输入框样式
        if hasattr(self, 'search_input'):
            from core.styles import BaseStyles
            input_style = BaseStyles.get_text_input_style(theme)
            self.search_input.setStyleSheet(input_style)
        
        # 更新下拉框样式
        if hasattr(self, 'version_filter'):
            from core.styles import BaseStyles
            combo_style = BaseStyles.get_combobox_style(theme)
            self.version_filter.setStyleSheet(combo_style)
        if hasattr(self, 'first_category_filter'):
            from core.styles import BaseStyles
            combo_style = BaseStyles.get_combobox_style(theme)
            self.first_category_filter.setStyleSheet(combo_style)
        if hasattr(self, 'second_category_filter'):
            from core.styles import BaseStyles
            combo_style = BaseStyles.get_combobox_style(theme)
            self.second_category_filter.setStyleSheet(combo_style)
        if hasattr(self, 'priority_filter'):
            from core.styles import BaseStyles
            combo_style = BaseStyles.get_combobox_style(theme)
            self.priority_filter.setStyleSheet(combo_style)
        if hasattr(self, 'hidden_filter'):
            from core.styles import BaseStyles
            combo_style = BaseStyles.get_combobox_style(theme)
            self.hidden_filter.setStyleSheet(combo_style)
        if hasattr(self, 'obtainable_filter'):
            from core.styles import BaseStyles
            combo_style = BaseStyles.get_combobox_style(theme)
            self.obtainable_filter.setStyleSheet(combo_style)
        
        # 更新表格样式
        if hasattr(self, 'table'):
            from core.styles import BaseStyles
            table_style = BaseStyles.get_text_input_style(theme)
            self.table.setStyleSheet(table_style)
    
    